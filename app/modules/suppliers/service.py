from sqlalchemy.orm import Session
from fastapi import HTTPException
from app.modules.suppliers import models, schemas


def get_supplier(db: Session, supplier_id: int, tenant_id: int):
    supplier = db.query(models.Supplier).filter(
        models.Supplier.id == supplier_id,
        models.Supplier.tenant_id == tenant_id
    ).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")
    return supplier


def get_suppliers(db: Session, tenant_id: int, skip: int = 0, limit: int = 100):
    return db.query(models.Supplier).filter(
        models.Supplier.tenant_id == tenant_id
    ).offset(skip).limit(limit).all()


def create_supplier(db: Session, supplier: schemas.SupplierCreate, tenant_id: int):
    db_supplier = models.Supplier(
        **supplier.model_dump(),
        tenant_id=tenant_id
    )
    db.add(db_supplier)
    db.commit()
    db.refresh(db_supplier)
    return db_supplier


def update_supplier(db: Session, supplier_id: int, supplier: schemas.SupplierUpdate, tenant_id: int):
    db_supplier = get_supplier(db, supplier_id, tenant_id)
    update_data = supplier.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_supplier, key, value)
    db.commit()
    db.refresh(db_supplier)
    return db_supplier


def delete_supplier(db: Session, supplier_id: int, tenant_id: int):
    db_supplier = get_supplier(db, supplier_id, tenant_id)
    db.delete(db_supplier)
    db.commit()
    return db_supplier


def export_to_excel(db: Session, tenant_id: int) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fornecedores"
    
    headers = [
        "ID", "Nome", "CNPJ", "Telefone 1", "Telefone 2", "E-mail"
    ]
    ws.append(headers)
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Skip limit to get all
    suppliers = db.query(models.Supplier).filter(
        models.Supplier.tenant_id == tenant_id
    ).all()
    
    for sup in suppliers:
        ws.append([
            sup.id, sup.name, sup.cnpj or "", 
            sup.phone_1 or "", sup.phone_2 or "", sup.email or ""
        ])
            
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
