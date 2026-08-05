from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from .repository import EmployeeRepository
from .schemas import EmployeeCreate, EmployeeUpdate, PublicBookingRequest

DEFAULT_WORKING_HOURS = {
    "0": {"is_open": False, "open": "08:00", "close": "12:00"},
    "1": {"is_open": True,  "open": "08:00", "close": "18:00"},
    "2": {"is_open": True,  "open": "08:00", "close": "18:00"},
    "3": {"is_open": True,  "open": "08:00", "close": "18:00"},
    "4": {"is_open": True,  "open": "08:00", "close": "18:00"},
    "5": {"is_open": True,  "open": "08:00", "close": "18:00"},
    "6": {"is_open": True,  "open": "08:00", "close": "18:00"}
}


class EmployeeService:
    def __init__(self):
        self.repository = EmployeeRepository()

    def create_employee(self, db: Session, tenant_id: int, data: EmployeeCreate):
        return self.repository.create(db, tenant_id, data)

    def get_employee(self, db: Session, tenant_id: int, employee_id: int):
        employee = self.repository.get_by_id(db, tenant_id, employee_id)
        if not employee:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Funcionário não encontrado.")
        return employee

    def list_employees(self, db: Session, tenant_id: int):
        return self.repository.list(db, tenant_id)

    def update_employee(self, db: Session, tenant_id: int, employee_id: int, data: EmployeeUpdate):
        employee = self.get_employee(db, tenant_id, employee_id)
        return self.repository.update(db, employee, data)

    def regenerate_token(self, db: Session, tenant_id: int, employee_id: int):
        import uuid
        employee = self.get_employee(db, tenant_id, employee_id)
        employee.schedule_token = uuid.uuid4().hex
        db.commit()
        db.refresh(employee)
        return employee

    def export_to_excel(self, db: Session, tenant_id: int) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Funcionários"
        
        headers = [
            "ID", "Nome", "Cargo", "Telefone", "E-mail", "Ativo"
        ]
        ws.append(headers)
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        employees = self.repository.list(db, tenant_id)
        
        role_map = {
            "groomer": "Tosador(a)",
            "bather": "Banhista",
            "vet": "Veterinário(a)",
            "salesperson": "Vendedor(a)",
            "receptionist": "Recepcionista",
            "driver": "Motorista",
            "other": "Outros"
        }
        
        for emp in employees:
            ws.append([
                emp.id, emp.name, role_map.get(emp.role, emp.role), 
                emp.phone or "", emp.email or "", 
                "Sim" if emp.is_active else "Não"
            ])
                
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def get_appointments_by_token(self, db: Session, token: str):
        from .models import Employee
        from app.modules.appointments.models import Appointment, AppointmentItem, AppointmentItemService

        employee = db.query(Employee).filter(Employee.schedule_token == token, Employee.is_active == True).first()
        if not employee:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agenda não encontrada ou funcionário inativo.")

        db_appointments = (
            db.query(Appointment)
            .join(AppointmentItem)
            .join(AppointmentItemService, AppointmentItemService.appointment_item_id == AppointmentItem.id)
            .filter(
                AppointmentItemService.employee_id == employee.id,
                Appointment.status != "canceled"
            )
            .order_by(Appointment.scheduled_at.asc())
            .all()
        )

        appointments_list = []
        for appt in db_appointments:
            items_list = []
            for item in appt.items:
                assigned_item_services = db.query(AppointmentItemService).filter(
                    AppointmentItemService.appointment_item_id == item.id,
                    AppointmentItemService.employee_id == employee.id
                ).all()
                assigned_service_ids = {ais.service_id for ais in assigned_item_services}

                if not assigned_service_ids:
                    continue

                services_list = []
                for svc in item.services:
                    if svc.id in assigned_service_ids:
                        services_list.append({
                            "id": svc.id,
                            "name": svc.name,
                            "duration_minutes": getattr(svc, "duration_minutes", None)
                        })

                if services_list:
                    items_list.append({
                        "id": item.id,
                        "pet": {
                            "id": item.pet.id,
                            "name": item.pet.name,
                            "species": item.pet.species,
                            "breed": item.pet.breed
                        },
                        "services": services_list
                    })

            if items_list:
                appointments_list.append({
                    "id": appt.id,
                    "scheduled_at": appt.scheduled_at,
                    "status": appt.status.value if hasattr(appt.status, "value") else str(appt.status),
                    "notes": appt.notes,
                    "client_name": appt.client.name,
                    "client_phone": appt.client.phone,
                    "items": items_list
                })

        from app.modules.tenants.models import Tenant
        tenant = db.query(Tenant).filter(Tenant.id == employee.tenant_id).first()
        petshop_name = tenant.name if tenant else "Petshop"

        return {
            "employee_name": employee.name,
            "employee_role": employee.role.value if hasattr(employee.role, "value") else str(employee.role),
            "petshop_name": petshop_name,
            "appointments": appointments_list
        }

    def get_public_booking_info(self, db: Session, token: str, appointment_id: int | None = None, sig: str | None = None):
        from .models import Employee
        from app.modules.appointments.models import Appointment, AppointmentItem, AppointmentItemService
        from app.modules.tenant_services.models import Service
        from app.modules.tenants.models import Tenant
        from app.config.settings import settings
        import hashlib
        from datetime import timedelta

        employee = db.query(Employee).filter(Employee.schedule_token == token, Employee.is_active == True).first()
        if not employee:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Profissional não encontrado.")

        tenant = db.query(Tenant).filter(Tenant.id == employee.tenant_id).first()
        if not tenant:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Petshop não encontrado.")

        reschedule_appointment = None
        if appointment_id and sig:
            # We must load the appointment first to check its values
            reschedule_appointment = db.query(Appointment).filter(
                Appointment.id == appointment_id,
                Appointment.tenant_id == tenant.id
            ).first()
            if not reschedule_appointment:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agendamento não encontrado.")
            
            from datetime import datetime, timezone
            from app.modules.appointments.models import AppointmentStatus

            # 1. Enforce status restrictions
            if reschedule_appointment.status not in (AppointmentStatus.PENDING, AppointmentStatus.CONFIRMED):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST, 
                    detail="Este agendamento já foi finalizado ou cancelado e não pode ser reagendado."
                )

            # 2. Verify signature using scheduled_at
            expected_sig = hashlib.sha256(f"{appointment_id}:{reschedule_appointment.scheduled_at.isoformat()}:{token}:{settings.JWT_SECRET_KEY}".encode("utf-8")).hexdigest()
            if sig != expected_sig:
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Assinatura de reagendamento inválida ou link expirado/já utilizado.")

        services = db.query(Service).filter(
            Service.tenant_id == employee.tenant_id,
            Service.is_active == True
        ).all()

        services_list = [{
            "id": s.id,
            "name": s.name,
            "description": s.description,
            "price_cents": s.price_cents,
            "duration_minutes": s.duration_minutes
        } for s in services]

        # Fetch ALL appointments for the entire petshop (Tenant level)
        db_appointments = (
            db.query(Appointment)
            .filter(
                Appointment.tenant_id == tenant.id,
                Appointment.status != "canceled"
            )
            .all()
        )

        busy_slots = []
        
        # Calculate capacity-blocked slots only if simultaneous limits are active (max_simultaneous_appointments is not None)
        if tenant.max_simultaneous_appointments is not None:
            intervals = []
            for appt in db_appointments:
                # Exclude the reschedule appointment itself from capacity calculations
                if reschedule_appointment and appt.id == reschedule_appointment.id:
                    continue
                    
                start = appt.scheduled_at
                duration = 0
                for item in appt.items:
                    for svc in item.services:
                        duration += svc.duration_minutes or 30
                end = start + timedelta(minutes=duration)
                intervals.append((start, end))

            # Sweep-line algorithm
            events = []
            for start, end in intervals:
                events.append((start, 1))
                events.append((end, -1))

            events.sort(key=lambda x: (x[0], x[1]))

            active_count = 0
            busy_start = None
            capacity = tenant.max_simultaneous_appointments

            for time_point, val in events:
                prev_count = active_count
                active_count += val

                if active_count >= capacity and prev_count < capacity:
                    busy_start = time_point
                elif active_count < capacity and prev_count >= capacity:
                    if busy_start and time_point > busy_start:
                        duration_min = int((time_point - busy_start).total_seconds() / 60)
                        busy_slots.append({
                            "scheduled_at": busy_start.isoformat(),
                            "duration_minutes": duration_min
                        })
                        busy_start = None

        reschedule_data = None
        if reschedule_appointment:
            services_info = []
            for item in reschedule_appointment.items:
                for svc in item.services:
                    services_info.append({
                        "id": svc.id,
                        "name": svc.name,
                        "price_cents": svc.price_cents,
                        "duration_minutes": svc.duration_minutes
                    })
            
            pet_name = reschedule_appointment.items[0].pet.name if (reschedule_appointment.items and reschedule_appointment.items[0].pet) else ""
            
            reschedule_data = {
                "id": reschedule_appointment.id,
                "client_name": reschedule_appointment.client.name,
                "client_phone": reschedule_appointment.client.phone,
                "pet_name": pet_name,
                "services": services_info,
                "notes": reschedule_appointment.notes
            }

        return {
            "employee_id": employee.id,
            "employee_name": employee.name,
            "employee_role": employee.role.value if hasattr(employee.role, "value") else str(employee.role),
            "petshop_name": tenant.name,
            "working_hours": tenant.working_hours or DEFAULT_WORKING_HOURS,
            "services": services_list,
            "busy_slots": busy_slots,
            "reschedule_appointment": reschedule_data
        }

    def create_public_booking(self, db: Session, token: str, data: PublicBookingRequest, appointment_id: int | None = None, sig: str | None = None):
        from .models import Employee
        from app.modules.appointments.models import Appointment, AppointmentItem, AppointmentStatus, AppointmentItemService
        from app.modules.tenant_services.models import Service
        from app.modules.clients.models import Client
        from app.modules.pets.models import Pet
        from app.modules.whatsapp.service import WhatsAppService
        from app.config.settings import settings
        from sqlalchemy import func
        import hashlib

        employee = db.query(Employee).filter(Employee.schedule_token == token, Employee.is_active == True).first()
        if not employee:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Profissional não encontrado.")

        # Handle reschedule logic
        if appointment_id:
            if not sig:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assinatura requerida para reagendamento.")
            
            appointment = db.query(Appointment).filter(
                Appointment.id == appointment_id,
                Appointment.tenant_id == employee.tenant_id
            ).first()
            if not appointment:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Agendamento não encontrado.")

            from datetime import datetime, timezone
            
            # 1. Enforce status restrictions
            if appointment.status not in (AppointmentStatus.PENDING, AppointmentStatus.CONFIRMED):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST, 
                    detail="Este agendamento já foi finalizado ou cancelado e não pode ser reagendado."
                )

            # 2. Verify signature using scheduled_at
            expected_sig = hashlib.sha256(f"{appointment_id}:{appointment.scheduled_at.isoformat()}:{token}:{settings.JWT_SECRET_KEY}".encode("utf-8")).hexdigest()
            if sig != expected_sig:
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Assinatura de reagendamento inválida ou link expirado/já utilizado.")

            # Update scheduled time & reset status
            appointment.scheduled_at = data.scheduled_at
            appointment.notes = data.notes
            appointment.status = AppointmentStatus.PENDING

            # Bind professional
            for item in appointment.items:
                item.employee_id = employee.id
                db.query(AppointmentItemService).filter(
                    AppointmentItemService.appointment_item_id == item.id
                ).update({"employee_id": employee.id}, synchronize_session=False)

            db.commit()

            try:
                WhatsAppService().send_appointment_confirmation(db, employee.tenant_id, appointment.id)
            except Exception:
                pass

            return {
                "success": True,
                "appointment_id": appointment.id,
                "message": "Agendamento reagendado com sucesso!"
            }

        # Original creation logic
        digits = "".join(c for c in data.client_phone if c.isdigit())
        suffix = digits[-9:] if len(digits) >= 9 else digits

        client = None
        if suffix:
            client = db.query(Client).filter(
                Client.tenant_id == employee.tenant_id,
                Client.phone.isnot(None),
                func.regexp_replace(Client.phone, '[^0-9]', '', 'g').like(f"%{suffix}")
            ).first()

        if not client:
            client = Client(
                tenant_id=employee.tenant_id,
                name=data.client_name,
                phone=data.client_phone,
                is_active=True
            )
            db.add(client)
            db.flush()

        pet = db.query(Pet).filter(
            Pet.client_id == client.id,
            Pet.tenant_id == employee.tenant_id,
            func.lower(Pet.name) == data.pet_name.strip().lower(),
            Pet.is_active == True
        ).first()

        if not pet:
            pet = Pet(
                tenant_id=employee.tenant_id,
                client_id=client.id,
                name=data.pet_name.strip(),
                species="dog",
                is_active=True
            )
            db.add(pet)
            db.flush()

        services = db.query(Service).filter(
            Service.id.in_(data.service_ids),
            Service.tenant_id == employee.tenant_id,
            Service.is_active == True
        ).all()

        if len(services) != len(data.service_ids):
            raise HTTPException(status_code=400, detail="Um ou mais serviços são inválidos.")

        appointment = Appointment(
            tenant_id=employee.tenant_id,
            client_id=client.id,
            scheduled_at=data.scheduled_at,
            notes=data.notes,
            status=AppointmentStatus.PENDING
        )
        db.add(appointment)
        db.flush()

        item = AppointmentItem(
            appointment_id=appointment.id,
            pet_id=pet.id,
            services=services,
            employee_id=employee.id
        )
        db.add(item)
        db.flush()

        db.query(AppointmentItemService).filter(
            AppointmentItemService.appointment_item_id == item.id
        ).update({"employee_id": employee.id}, synchronize_session=False)
        
        db.commit()

        try:
            WhatsAppService().send_appointment_confirmation(db, employee.tenant_id, appointment.id)
        except Exception:
            pass

        return {
            "success": True,
            "appointment_id": appointment.id,
            "message": "Agendamento realizado com sucesso!"
        }

    def delete_employee(self, db: Session, tenant_id: int, employee_id: int):
        employee = self.get_employee(db, tenant_id, employee_id)
        self.repository.delete(db, employee)
