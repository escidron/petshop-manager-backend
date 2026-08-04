from datetime import date, datetime
import io
from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from .repository import ClientRepository
from .schemas import ClientCreate, ClientUpdate
from app.modules.pets.repository import PetRepository
from app.modules.pets.schemas import PetCreate


class ClientService:
    def __init__(self):
        self.repository = ClientRepository()
        self.pet_repository = PetRepository()

    def create_client(
        self, db: Session, tenant_id: int, data: ClientCreate
    ):
        return self.repository.create(db, tenant_id, data)

    def get_client(
        self, db: Session, tenant_id: int, client_id: int
    ):
        client = self.repository.get_by_id(
            db, tenant_id, client_id
        )
        if not client:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Cliente não encontrado",
            )
        return client

    def list_clients(self, db: Session, tenant_id: int):
        return self.repository.list(db, tenant_id)

    def update_client(
        self,
        db: Session,
        tenant_id: int,
        client_id: int,
        data: ClientUpdate,
    ):
        client = self.get_client(db, tenant_id, client_id)
        return self.repository.update(db, client, data)

    def delete_client(
        self, db: Session, tenant_id: int, client_id: int
    ):
        client = self.get_client(db, tenant_id, client_id)
        self.repository.delete(db, client)

    def generate_import_template_excel(self) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes e Pets"
        
        # Add headers
        headers = [
            "cliente_nome *", "cliente_telefone *", "cliente_email", "cliente_documento", "cliente_data_nascimento", 
            "cliente_cep", "cliente_logradouro", "cliente_numero", "cliente_complemento", "cliente_bairro", "cliente_cidade", "cliente_estado",
            "pet_nome *", "pet_especie *", "pet_raca", "pet_tipo_pelagem", "pet_cor_pelagem", "pet_genero", "pet_porte", "pet_castrado", "pet_data_nascimento", "pet_idade_aproximada", "pet_unidade_idade", "pet_observacoes"
        ]
        ws.append(headers)
        
        # Style headers (Client is light blue 1976D2, Pet is teal 00796B, Mandatory is Red D32F2F)
        font_style = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        client_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        pet_fill = PatternFill(start_color="00796B", end_color="00796B", fill_type="solid")
        mandatory_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3")
        )
        
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_style
            cell.alignment = center_align
            cell.border = thin_border
            # Apply color based on group
            if "*" in headers[col_idx - 1]:
                cell.fill = mandatory_fill
            elif col_idx <= 12:
                cell.fill = client_fill
            else:
                cell.fill = pet_fill
                
        # Add examples
        examples = [
            [
                "João Silva", "(11) 99999-8888", "joao@email.com", "123.456.789-00", "15/08/1985", 
                "01311-200", "Avenida Paulista", "1000", "Apto 12", "Bela Vista", "São Paulo", "SP",
                "Rex", "Canino", "Golden Retriever", "Longa", "Caramelo", "Macho", "G", "Não", "01/05/2020", "", "", "Alérgico a medicamentos"
            ],
            [
                "João Silva", "(11) 99999-8888", "joao@email.com", "123.456.789-00", "15/08/1985", 
                "01311-200", "Avenida Paulista", "1000", "Apto 12", "Bela Vista", "São Paulo", "SP",
                "Mia", "Felino", "Siamês", "Curta", "Bicolor", "Fêmea", "P", "Sim", "", "2", "Anos", "Muito dócil e assustada"
            ],
            [
                "Maria Souza", "(11) 88888-7777", "maria@email.com", "987.654.321-11", "20/10/1990", 
                "04533-010", "", "500", "Fundos", "", "", "",
                "Fred", "Exótico", "Jabuti", "Sem pelo", "Marrom", "Macho", "PP", "Não", "01/01/2018", "", "", "Vive em terrário"
            ]
        ]
        
        for ex in examples:
            ws.append(ex)
            
        # Format rows up to 1000 (pre-styles cells so they format cleanly when typed)
        for r_idx in range(2, 1001):
            ws.row_dimensions[r_idx].height = 20
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                
                # Alignment
                if col_idx in [1, 3, 7, 9, 10, 11, 13, 15, 16, 17, 24]:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                    
                # Format E (5) and U (21) as custom date mask
                if col_idx in [5, 21]:
                    cell.number_format = '00\\/00\\/0000'
                # Format B (2) as custom Phone mask
                elif col_idx == 2:
                    cell.number_format = '\\(00\\)\\ 00000\\-0000'
                # Format D (4) as custom CPF/CNPJ conditional mask
                elif col_idx == 4:
                    cell.number_format = '[>99999999999]00\\.000\\.000\\/0000\\-00;000\\.000\\.000\\-00'
                # Format F (6) as custom CEP mask
                elif col_idx == 6:
                    cell.number_format = '00000\\-000'
                    
        # Add validation rules
        # 1. Species (Column N - col 14)
        dv_species = DataValidation(type="list", formula1='"Canino,Felino,Exótico"', allow_blank=True)
        dv_species.error = 'Escolha uma espécie da lista (Canino, Felino ou Exótico)'
        dv_species.errorTitle = 'Espécie Inválida'
        dv_species.prompt = 'Selecione a espécie (Canino, Felino ou Exótico)'
        dv_species.promptTitle = 'Espécie'
        ws.add_data_validation(dv_species)
        dv_species.add("N2:N1000")
        
        # 2. Coat Type (Column P - col 16)
        dv_coat_type = DataValidation(type="list", formula1='"Curta,Média,Longa,Dupla camada,Sem pelo"', allow_blank=True)
        dv_coat_type.error = 'Escolha um tipo de pelagem da lista'
        dv_coat_type.errorTitle = 'Pelagem Inválida'
        dv_coat_type.prompt = 'Selecione o tipo de pelagem'
        dv_coat_type.promptTitle = 'Pelagem'
        ws.add_data_validation(dv_coat_type)
        dv_coat_type.add("P2:P1000")
        
        # 3. Coat Color (Column Q - col 17)
        dv_coat_color = DataValidation(type="list", formula1='"Branco,Preto,Cinza,Marrom,Caramelo,Bege,Dourado,Creme,Chocolate,Canela,Fulvo,Rajado,Tigrado,Malhado,Mesclado,Tricolor,Bicolor,Pardo,Amarelo,Vermelho"', allow_blank=True)
        dv_coat_color.error = 'Escolha uma cor da lista'
        dv_coat_color.errorTitle = 'Cor Inválida'
        dv_coat_color.prompt = 'Selecione a cor principal'
        dv_coat_color.promptTitle = 'Cor'
        ws.add_data_validation(dv_coat_color)
        dv_coat_color.add("Q2:Q1000")
        
        # 4. Gender (Column R - col 18)
        dv_gender = DataValidation(type="list", formula1='"Macho,Fêmea,Desconhecido"', allow_blank=True)
        dv_gender.error = 'Escolha um gênero da lista'
        dv_gender.errorTitle = 'Gênero Inválido'
        dv_gender.prompt = 'Selecione o gênero'
        dv_gender.promptTitle = 'Gênero'
        ws.add_data_validation(dv_gender)
        dv_gender.add("R2:R1000")
        
        # 4. Size (Column S - col 19)
        dv_size = DataValidation(type="list", formula1='"PP,P,M,G,GG"', allow_blank=True)
        dv_size.error = 'Escolha um porte da lista'
        dv_size.errorTitle = 'Porte Inválido'
        dv_size.prompt = 'Selecione o porte'
        dv_size.promptTitle = 'Porte'
        ws.add_data_validation(dv_size)
        dv_size.add("S2:S1000")
        
        # 5. Neutered (Column T - col 20)
        dv_neutered = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
        dv_neutered.error = 'Escolha Sim ou Não'
        dv_neutered.errorTitle = 'Opção Inválida'
        dv_neutered.prompt = 'O animal é castrado?'
        dv_neutered.promptTitle = 'Castrado'
        ws.add_data_validation(dv_neutered)
        dv_neutered.add("T2:T1000")

        # 6. Age Unit (Column W - col 23)
        dv_age_unit = DataValidation(type="list", formula1='"Meses,Anos"', allow_blank=True)
        dv_age_unit.error = 'Escolha Meses ou Anos'
        dv_age_unit.errorTitle = 'Unidade Inválida'
        dv_age_unit.prompt = 'Meses ou Anos?'
        dv_age_unit.promptTitle = 'Unidade'
        ws.add_data_validation(dv_age_unit)
        dv_age_unit.add("W2:W1000")
        
        # Column Widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)
            
        ws.views.sheetView[0].showGridLines = True
        
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    async def import_clients_from_excel(self, db: Session, tenant_id: int, file_content: bytes, progress_callback=None):
        import openpyxl
        import io
        from datetime import date
        from app.modules.clients.models import Client

        if not file_content:
            return {"imported": 0, "errors": ["Arquivo de planilha vazio"]}

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return {"imported": 0, "errors": [f"Erro ao ler arquivo Excel: {str(e)}"]}

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"imported": 0, "errors": ["A planilha está vazia"]}

        raw_headers = rows[0]
        headers = []
        for h in raw_headers:
            if h is None:
                headers.append("")
            else:
                headers.append(str(h).strip().lower().replace("*", "").strip())

        required_cols = ["cliente_nome", "cliente_telefone"]
        missing_cols = [col for col in required_cols if col not in headers]
        if missing_cols:
            return {
                "imported": 0,
                "errors": [f"Colunas obrigatórias ausentes na planilha: {', '.join(missing_cols)}"],
            }

        imported_count = 0
        errors = []

        # ---- Helpers ----
        def clean_str(val):
            if val is None or str(val).strip() == "":
                return None
            val_str = str(val).strip()
            if isinstance(val, float) and val.is_integer():
                return str(int(val))
            return val_str

        def clean_num(val):
            if val is None or str(val).strip() == "":
                return None
            return "".join(filter(str.isdigit, str(val)))

        def parse_date(val):
            if val is None or str(val).strip() == "":
                return None
            if hasattr(val, "date"):
                return val.date()
            if isinstance(val, date):
                return val
            try:
                if isinstance(val, (int, float)):
                    cleaned = str(int(val))
                else:
                    cleaned = str(val).strip().replace("/", "").replace("-", "")
                    if "." in cleaned:
                        cleaned = cleaned.split(".")[0]
                if len(cleaned) == 7 and cleaned.isdigit():
                    cleaned = "0" + cleaned
                if len(cleaned) == 8 and cleaned.isdigit():
                    return date(int(cleaned[4:8]), int(cleaned[2:4]), int(cleaned[0:2]))
                cleaned_orig = str(val).strip()
                if "/" in cleaned_orig:
                    parts = cleaned_orig.split("/")
                    if len(parts) == 3:
                        return date(int(parts[2]), int(parts[1]), int(parts[0]))
                elif "-" in cleaned_orig:
                    parts = cleaned_orig.split("-")
                    if len(parts) == 3:
                        if len(parts[0]) == 4:
                            return date(int(parts[0]), int(parts[1]), int(parts[2]))
                        else:
                            return date(int(parts[2]), int(parts[1]), int(parts[0]))
            except Exception:
                pass
            return None

        def parse_species(val):
            if not val:
                return None
            cleaned = str(val).strip().lower()
            if "canino" in cleaned or "cão" in cleaned or "cao" in cleaned or "cachorro" in cleaned:
                return "Canino"
            if "felino" in cleaned or "gato" in cleaned:
                return "Felino"
            if "exotico" in cleaned or "exótico" in cleaned or "exoticos" in cleaned:
                return "Exoticos"
            return val

        def parse_gender(val):
            if not val:
                return "unknown"
            cleaned = str(val).strip().lower()
            if "macho" in cleaned:
                return "male"
            if "fêmea" in cleaned or "femea" in cleaned:
                return "female"
            return "unknown"

        def parse_age_unit(val):
            if not val:
                return None
            cleaned = str(val).strip().lower()
            if "meses" in cleaned or "mês" in cleaned or "mes" in cleaned:
                return "months"
            if "anos" in cleaned or "ano" in cleaned:
                return "years"
            return None

        def parse_coat_type(val):
            if not val:
                return None
            cleaned = str(val).strip().lower()
            if "curta" in cleaned:
                return "short"
            if "média" in cleaned or "media" in cleaned:
                return "medium"
            if "longa" in cleaned:
                return "long"
            if "dupla" in cleaned:
                return "double"
            if "sem pelo" in cleaned:
                return "hairless"
            return val

        def parse_neutered(val):
            if not val:
                return None
            cleaned = str(val).strip().lower()
            if "sim" in cleaned:
                return True
            if "não" in cleaned or "nao" in cleaned:
                return False
            return None

        def parse_size(val):
            if not val:
                return None
            cleaned = str(val).strip().lower()
            if "mini" in cleaned or "pp" in cleaned or "micro" in cleaned:
                return "PP"
            elif "pequeno" in cleaned or " p " in f" {cleaned} " or cleaned == "p":
                return "P"
            elif "médio" in cleaned or "medio" in cleaned or " m " in f" {cleaned} " or cleaned == "m":
                return "M"
            elif "grande" in cleaned or " g " in f" {cleaned} " or cleaned == "g":
                return "G"
            elif "gigante" in cleaned or "gg" in cleaned or "extra" in cleaned:
                return "GG"
            return str(val).strip()[:5]


        # Filter completely empty rows early
        non_empty_rows = [
            rv for rv in rows[1:]
            if any(v is not None and str(v).strip() != "" for v in rv)
        ]

        # ------------------------------------------------------------------ #
        #  OPTIMISATION 1: Pre-fetch all unique CEPs in parallel              #
        # ------------------------------------------------------------------ #
        cep_col_idx = headers.index("cliente_cep") if "cliente_cep" in headers else None
        cep_cache: dict[str, dict] = {}

        if cep_col_idx is not None:
            import asyncio
            from app.modules.address.service import AddressService

            unique_ceps: set[str] = set()
            for row_values in non_empty_rows:
                raw_cep = row_values[cep_col_idx] if cep_col_idx < len(row_values) else None
                cep_digits = clean_num(raw_cep)
                if cep_digits:
                    if len(cep_digits) == 7:
                        cep_digits = "0" + cep_digits
                    if len(cep_digits) == 8:
                        unique_ceps.add(cep_digits)

            async def _fetch_cep(cep: str):
                try:
                    return cep, await AddressService.fetch_by_cep(cep)
                except Exception:
                    return cep, None

            fetched = await asyncio.gather(*[_fetch_cep(c) for c in unique_ceps])
            for cep_key, addr_data in fetched:
                if addr_data:
                    cep_cache[cep_key] = addr_data

        # ------------------------------------------------------------------ #
        #  OPTIMISATION 2: Load existing clients/pets into memory (1 query)   #
        # ------------------------------------------------------------------ #
        existing_clients_q = db.query(Client).filter(Client.tenant_id == tenant_id).all()
        existing_by_name_phone: dict[tuple, Client] = {}
        existing_by_document: dict[str, Client] = {}
        for ec in existing_clients_q:
            norm_name = ec.name.lower().strip() if ec.name else ""
            norm_phone = "".join(filter(str.isdigit, ec.phone)) if ec.phone else ""
            existing_by_name_phone[(norm_name, norm_phone)] = ec
            if ec.document:
                existing_by_document[ec.document] = ec

        from app.modules.pets.models import Pet
        existing_pets_q = db.query(Pet).filter(Pet.tenant_id == tenant_id).all()
        existing_pets: set[tuple] = {
            (ep.client_id, (ep.name or "").lower(), (ep.species or "").lower())
            for ep in existing_pets_q
        }

        # ------------------------------------------------------------------ #
        #  MAIN LOOP — collect objects and flush every CHUNK_SIZE rows        #
        # ------------------------------------------------------------------ #
        CHUNK_SIZE = 200
        total_rows = len(non_empty_rows)
        session_client_cache: dict = {}
        new_clients: list[Client] = []
        new_pets: list[Pet] = []
        created_clients = 0
        updated_clients = 0
        processed_count = 0

        # Inform caller of total upfront so frontend can show 'X de Y'
        if progress_callback and total_rows > 0:
            progress_callback(0, total_rows, 0, 0)

        def _flush_chunk():
            nonlocal new_clients, new_pets
            if new_clients:
                db.add_all(new_clients)
                db.flush()  # assigns DB IDs without committing yet
                new_clients = []
            # Assign client_id to pets that were created alongside new clients
            for pet_obj in new_pets:
                pending_client = getattr(pet_obj, "_pending_client", None)
                if pending_client is not None:
                    pet_obj.client_id = pending_client.id
                    del pet_obj._pending_client
            if new_pets:
                db.add_all(new_pets)
                db.flush()
                new_pets = []
            db.commit()

        for row_idx, row_values in enumerate(non_empty_rows, start=2):
            try:
                row = {h: val for h, val in zip(headers, row_values) if h}

                # 1. Client validation
                name = clean_str(row.get("cliente_nome"))
                phone = clean_str(row.get("cliente_telefone"))
                if not name:
                    raise ValueError("Nome do cliente é obrigatório")
                if not phone:
                    raise ValueError("Telefone do cliente é obrigatório")

                email = clean_str(row.get("cliente_email"))

                document = clean_num(row.get("cliente_documento"))
                if document:
                    if len(document) == 10:
                        document = "0" + document
                    elif len(document) == 13:
                        document = "0" + document

                document_type = None
                if document:
                    if len(document) == 11:
                        document_type = "CPF"
                    elif len(document) == 14:
                        document_type = "CNPJ"

                birth_date = parse_date(row.get("cliente_data_nascimento"))

                cep = clean_num(row.get("cliente_cep"))
                if cep and len(cep) == 7:
                    cep = "0" + cep

                street = clean_str(row.get("cliente_logradouro"))
                number = clean_str(row.get("cliente_numero"))
                complement = clean_str(row.get("cliente_complemento"))
                neighborhood = clean_str(row.get("cliente_bairro"))
                city = clean_str(row.get("cliente_cidade"))
                state = clean_str(row.get("cliente_estado"))

                # Use pre-fetched CEP cache (no HTTP call per row)
                if cep and cep in cep_cache and not (street and neighborhood and city and state):
                    addr_data = cep_cache[cep]
                    street = street or addr_data.get("street")
                    neighborhood = neighborhood or addr_data.get("neighborhood")
                    city = city or addr_data.get("city")
                    state = state or addr_data.get("state")

                # Resolve client — session cache → memory set → create or update
                norm_name = name.lower().strip()
                norm_phone = "".join(filter(str.isdigit, phone))
                cache_key = document if document else (norm_name, norm_phone)

                client = session_client_cache.get(cache_key)
                if not client and document:
                    client = existing_by_document.get(document)
                if not client:
                    client = existing_by_name_phone.get((norm_name, norm_phone))

                is_new_client = client is None

                if is_new_client:
                    # CREATE
                    client = Client(
                        tenant_id=tenant_id,
                        name=name,
                        phone=phone,
                        email=email,
                        document=document,
                        document_type=document_type,
                        birth_date=birth_date,
                        cep=cep,
                        street=street,
                        number=number,
                        complement=complement,
                        neighborhood=neighborhood,
                        city=city,
                        state=state,
                        is_active=True,
                    )
                    new_clients.append(client)
                    if document:
                        existing_by_document[document] = client
                    existing_by_name_phone[(norm_name, norm_phone)] = client
                    created_clients += 1
                else:
                    # UPDATE — only overwrite fields that have a value in the file
                    _changed = False
                    for attr, val in [
                        ("email", email), ("document", document), ("document_type", document_type),
                        ("birth_date", birth_date), ("cep", cep), ("street", street),
                        ("number", number), ("complement", complement),
                        ("neighborhood", neighborhood), ("city", city), ("state", state),
                    ]:
                        if val is not None and getattr(client, attr, None) != val:
                            setattr(client, attr, val)
                            _changed = True
                    if _changed:
                        updated_clients += 1

                session_client_cache[cache_key] = client

                # 2. Pet validation & creation
                pet_name = clean_str(row.get("pet_nome"))
                if not pet_name:
                    raise ValueError("Nome do pet é obrigatório")

                pet_species_raw = clean_str(row.get("pet_especie"))
                if not pet_species_raw:
                    raise ValueError("Espécie do pet é obrigatória")

                pet_species = parse_species(pet_species_raw)
                pet_breed = clean_str(row.get("pet_raca"))
                pet_coat_type = parse_coat_type(row.get("pet_tipo_pelagem"))
                pet_coat_color = clean_str(row.get("pet_cor_pelagem"))
                pet_gender = parse_gender(clean_str(row.get("pet_genero")))
                pet_size = parse_size(row.get("pet_porte"))
                pet_neutered = parse_neutered(clean_str(row.get("pet_castrado")))

                pet_birth = parse_date(row.get("pet_data_nascimento"))
                pet_age_str = clean_num(row.get("pet_idade_aproximada"))
                pet_age = int(pet_age_str) if pet_age_str else None
                pet_age_unit = parse_age_unit(row.get("pet_unidade_idade"))
                pet_notes = clean_str(row.get("pet_observacoes"))

                is_new_client_obj = client in new_clients
                if is_new_client_obj:
                    pet_mem_key = (id(client), pet_name.lower(), (pet_species or "").lower())
                else:
                    pet_mem_key = (client.id, pet_name.lower(), (pet_species or "").lower())

                if pet_mem_key not in existing_pets:
                    # CREATE pet
                    pet_obj = Pet(
                        tenant_id=tenant_id,
                        client_id=None,
                        name=pet_name,
                        species=pet_species,
                        breed=pet_breed,
                        coat_type=pet_coat_type,
                        coat_color=pet_coat_color,
                        gender=pet_gender,
                        size=pet_size,
                        is_neutered=pet_neutered,
                        birth_date=pet_birth,
                        age=pet_age,
                        age_unit=pet_age_unit,
                        notes=pet_notes,
                        is_active=True,
                    )
                    if is_new_client_obj:
                        pet_obj._pending_client = client
                    else:
                        pet_obj.client_id = client.id
                    new_pets.append(pet_obj)
                    existing_pets.add(pet_mem_key)
                else:
                    # UPDATE pet — overwrite non-empty fields
                    if not is_new_client_obj:
                        existing_pet = db.query(Pet).filter(
                            Pet.tenant_id == tenant_id,
                            Pet.client_id == client.id,
                            Pet.name == pet_name,
                            Pet.species == pet_species,
                        ).first()
                        if existing_pet:
                            for attr, val in [
                                ("breed", pet_breed), ("coat_type", pet_coat_type),
                                ("coat_color", pet_coat_color), ("gender", pet_gender),
                                ("size", pet_size), ("is_neutered", pet_neutered),
                                ("birth_date", pet_birth), ("age", pet_age),
                                ("age_unit", pet_age_unit), ("notes", pet_notes),
                            ]:
                                if val is not None and getattr(existing_pet, attr, None) != val:
                                    setattr(existing_pet, attr, val)

                processed_count += 1

            except Exception as e:
                errors.append(f"Linha {row_idx}: {str(e)}")

            # Flush chunk every CHUNK_SIZE rows and report progress
            if (row_idx - 1) % CHUNK_SIZE == 0:
                _flush_chunk()
                if progress_callback and total_rows > 0:
                    pct = min(int((row_idx - 1) / total_rows * 100), 99)
                    progress_callback(row_idx - 1, total_rows, processed_count, pct)

        # Final flush for remaining items
        _flush_chunk()

        return {
            "created": created_clients,
            "updated": updated_clients,
            "total": total_rows,
            "errors": errors,
        }

    async def import_clients_from_excel_background(
        self,
        job_id: str,
        tenant_id: int,
        file_content: bytes,
    ) -> None:
        """
        Runs import_clients_from_excel as a background task, reporting real-time
        progress (imported count + percentage) after every chunk flush.
        """
        from app.modules.clients.import_jobs import update_job
        from app.config.database import SessionLocal

        db = SessionLocal()
        try:
            update_job(job_id, status="running", progress=0)

            def _on_progress(processed: int, total: int, count: int, pct: int):
                update_job(job_id, progress=pct, imported=count, total=total)

            result = await self.import_clients_from_excel(
                db, tenant_id, file_content,
                progress_callback=_on_progress,
            )
            update_job(
                job_id,
                status="done",
                progress=100,
                created=result["created"],
                updated=result["updated"],
                total=result.get("total", 0),
                errors=result["errors"],
            )
        except Exception as e:
            update_job(job_id, status="error", errors=[str(e)])
        finally:
            db.close()

    def export_to_excel(self, db: Session, tenant_id: int) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO

        wb = openpyxl.Workbook()
        
        # Sheet 1: Clientes
        ws_clients = wb.active
        ws_clients.title = "Clientes"
        
        client_headers = [
            "ID", "Nome", "CPF/CNPJ", "Email", "Telefone", "Telefone 2", 
            "Data Nascimento", "CEP", "Logradouro", "Número", "Complemento", "Bairro", "Cidade", "Estado", "Ativo"
        ]
        ws_clients.append(client_headers)
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        client_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        
        for col_idx in range(1, len(client_headers) + 1):
            cell = ws_clients.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = client_fill

        clients = self.repository.list(db, tenant_id)
        
        for c in clients:
            ws_clients.append([
                c.id, c.name, c.document or "", c.email or "", c.phone or "", c.phone_secondary or "",
                c.birth_date.strftime("%d/%m/%Y") if c.birth_date else "",
                c.cep or "", c.street or "", c.number or "", c.complement or "", c.neighborhood or "", c.city or "", c.state or "",
                "Sim" if c.is_active else "Não"
            ])
            
        # Sheet 2: Pets
        ws_pets = wb.create_sheet("Pets")
        pet_headers = [
            "ID", "ID Cliente", "Nome Cliente", "Nome Pet", "Espécie", "Raça", "Porte", 
            "Tipo Pelagem", "Cor", "Gênero", "Castrado", "Data Nascimento", "Idade", "Unidade Idade", "Ativo", "Falecido", "Observações"
        ]
        ws_pets.append(pet_headers)
        
        pet_fill = PatternFill(start_color="00796B", end_color="00796B", fill_type="solid")
        for col_idx in range(1, len(pet_headers) + 1):
            cell = ws_pets.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = pet_fill
            
        coat_map = {
            "short": "Curta",
            "medium": "Média",
            "long": "Longa",
            "hairless": "Sem Pelo",
            "double": "Dupla",
            "curly": "Encaracolada"
        }
        
        gender_map = {
            "male": "Macho",
            "female": "Fêmea",
            "unknown": "Desconhecido"
        }
        
        age_unit_map = {
            "months": "Meses",
            "years": "Anos"
        }

        species_map = {
            "dog": "Canino",
            "cat": "Felino",
            "bird": "Ave",
            "other": "Outros"
        }
            
        for c in clients:
            for p in c.pets:
                coat_val = coat_map.get(p.coat_type, p.coat_type) if p.coat_type else ""
                gender_val = gender_map.get(p.gender, p.gender) if p.gender else ""
                age_unit_val = age_unit_map.get(p.age_unit, p.age_unit) if p.age_unit else ""
                species_val = species_map.get(p.species, p.species) if p.species else ""
                
                ws_pets.append([
                    p.id, c.id, c.name, p.name, species_val, p.breed or "", p.size or "",
                    coat_val, p.coat_color or "", gender_val, "Sim" if p.is_neutered else "Não",
                    p.birth_date.strftime("%d/%m/%Y") if p.birth_date else "",
                    p.age or "", age_unit_val, "Sim" if p.is_active else "Não", "Sim" if p.is_deceased else "Não", p.notes or ""
                ])
                
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

