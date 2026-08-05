from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.modules.auth.token import hash_password, verify_password
from app.modules.users.schemas import UserCreate, UserUpdate
from .repository import UserRepository


class UserService:
    def __init__(self):
        self.repository = UserRepository()

    def create(self, db: Session, data: UserCreate):
        user_data = data.model_dump()  # Pydantic v2
        user_data["password"] = hash_password(user_data["password"])
        
        # Evitar escalada de privilégios (Mass Assignment): não permite criar com role 'admin'
        if user_data.get("role") == "admin":
            user_data["role"] = "owner"
            
        return self.repository.create(db, user_data)

    def get_user(self, db: Session, user_id: int):
        user = self.repository.get_user(db, user_id)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Usuário não encontrado",
            )
        return user

    def change_password(self, db: Session, user_id: int, current_password: str, new_password: str):
        from app.modules.users.models import User
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuário não encontrado")
        if not verify_password(current_password, user.password):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Senha atual incorreta")
        user.password = hash_password(new_password)
        db.commit()

    def update_user(self, db: Session, user_id: int, data: UserUpdate):
        from app.modules.users.models import User
        from app.modules.auth.service import AuthService
        
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuário não encontrado")

        email_changed = False
        if data.name is not None:
            user.name = data.name
        if data.phone is not None:
            user.phone = data.phone
        if data.email is not None and data.email != user.email:
            existing = db.query(User).filter(User.email == data.email).first()
            if existing:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="E-mail já cadastrado por outro usuário")
            user.email = data.email
            user.email_verified = False
            email_changed = True

        db.commit()

        if email_changed:
            try:
                auth_service = AuthService()
                auth_service.resend_verification_email(db, user.id)
            except Exception as e:
                print(f"Error sending verification code after email change: {e}")

        return user

    def export_to_excel(self, db: Session, tenant_id: int) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        from app.modules.users.models import User, TenantUser

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Funcionários"
        
        headers = [
            "ID", "Nome", "E-mail", "Telefone", "Cargo", "Data Cadastro", "Ativo"
        ]
        ws.append(headers)
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        # Join TenantUser and User
        users = db.query(TenantUser, User).join(User, TenantUser.user_id == User.id).filter(
            TenantUser.tenant_id == tenant_id
        ).all()
        
        role_map = {
            "owner": "Proprietário",
            "employee": "Funcionário"
        }
        
        for t_user, user in users:
            role_val = role_map.get(t_user.role, t_user.role)
            ws.append([
                user.id, user.name, user.email, user.phone or "", 
                role_val, 
                t_user.created_at.strftime("%d/%m/%Y") if t_user.created_at else "",
                "Sim" if t_user.active else "Não"
            ])
                
        out = BytesIO()
        wb.save(out)
        return out.getvalue()
