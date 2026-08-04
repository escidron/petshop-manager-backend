from fastapi import HTTPException
from sqlalchemy.orm import Session


from app.modules.tenant_services.constants import DEFAULT_SERVICES
from app.modules.tenant_services.schemas import ServiceCreate

from .repository import ServiceRepository

class ServiceService:
    def __init__(self):
        self.repo = ServiceRepository()

    def create(self, db, tenant_id, data: ServiceCreate):
        # Validar duplicidade
        existing = self.repo.get_by_attributes(
            db,
            tenant_id=tenant_id,
            name=data.name,
            species=data.species,
            size=data.size,
            coat_type=data.coat_type
        )
        if existing:
                raise HTTPException(
                    status_code=400,
                    detail="Já existe um serviço com este nome e as mesmas variações (Espécie, Porte, Pelagem). Para preços diferentes, especifique as variações correspondentes."
                )
            
        return self.repo.create(db, tenant_id, data)

    def list(self, db, tenant_id):
        return self.repo.list(db, tenant_id)

    def get(self, db, tenant_id, service_id):
        service = self.repo.get_by_id(
            db, tenant_id, service_id
        )
        if not service:
            raise HTTPException(404, "Serviço não encontrado")
        return service

    def update(self, db, tenant_id, service_id, data):
        service = self.get(db, tenant_id, service_id)
        
        # Se algum campo identificador mudar, validar duplicidade
        if any(v is not None for v in [data.name, data.species, data.size, data.coat_type]):
            new_name = data.name if data.name is not None else service.name
            new_species = data.species if data.species is not None else service.species
            new_size = data.size if data.size is not None else service.size
            new_coat = data.coat_type if data.coat_type is not None else service.coat_type
            
            existing = self.repo.get_by_attributes(
                db,
                tenant_id=tenant_id,
                name=new_name,
                species=new_species,
                size=new_size,
                coat_type=new_coat
            )
            
            if existing and existing.id != service_id:
                raise HTTPException(
                    status_code=400,
                    detail="Já existe um serviço cadastrado com este nome e variações."
                )

        return self.repo.update(db, service, data)

    def delete(self, db, tenant_id, service_id):
        service = self.get(db, tenant_id, service_id)
        self.repo.delete(db, service)

    def create_default_services(
        self,
        db,
        tenant_id: int,
    ):
        for service in DEFAULT_SERVICES:
            data = ServiceCreate(**service)
            self.repo.create(
                db,
                tenant_id=tenant_id,
                data=data,
            )

    def generate_import_template_excel(self) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Serviços"
        
        # Add headers
        headers = [
            "servico *", "preco *", "especie", "porte", "pelagem", "duracao_minutos", "descricao"
        ]
        ws.append(headers)
        
        # Style headers (Segoe UI, Blue color theme)
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        required_header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3")
        )
        
        # Format header row
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            header_name = headers[col_idx - 1]
            if "*" in header_name:
                cell.fill = required_header_fill
            else:
                cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        base_services = ["Banho", "Tosa", "Banho & Tosa"]
        species_list = ["Canino", "Felino"]
        sizes_list = ["PP", "P", "M", "G", "GG"]
        coat_types = [None, "Curta", "Média", "Longa", "Dupla"]
        size_descriptions = {
            "PP": "porte muito pequeno (PP)",
            "P": "porte pequeno (P)",
            "M": "porte médio (M)",
            "G": "porte grande (G)",
            "GG": "porte muito grande (GG)"
        }

        # Append predefined rows
        for service in base_services:
            for species in species_list:
                for size in sizes_list:
                    description = f"{service} para {species.lower()} de {size_descriptions[size]}"
                    if species == "Canino":
                        for coat in coat_types:
                            coat_desc = f" com pelagem {coat.lower()}" if coat else ""
                            ws.append([
                                service,
                                "", # preco *
                                species,
                                size,
                                coat if coat else "",
                                "", # duracao_minutos
                                f"{description}{coat_desc}"
                            ])
                    else:
                        ws.append([
                            service,
                            "", # preco *
                            species,
                            size,
                            "",
                            "", # duracao_minutos
                            description
                        ])

        # Format rows (starting at row 2 up to ws.max_row)
        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                
                # Alignments & formatting based on column
                if col_idx in [1, 7]: # service, description -> left
                    cell.alignment = left_align
                else: # species, size, coat_type, price, duration -> center
                    cell.alignment = center_align
                    
                # Number formats
                if col_idx == 2: # price
                    cell.number_format = "#,##0.00"
                elif col_idx == 6: # duration
                    cell.number_format = "#,##0"

        # Data Validations (Dropdowns)
        # Species (Column 3)
        dv_species = DataValidation(type="list", formula1=f'"{",".join(species_list)}"', allow_blank=True)
        dv_species.error = 'Escolha uma espécie da lista (Canino ou Felino)'
        dv_species.errorTitle = 'Espécie Inválida'
        dv_species.prompt = 'Selecione a espécie atendida'
        dv_species.promptTitle = 'Espécie'
        ws.add_data_validation(dv_species)
        dv_species.add(f"C2:C{ws.max_row + 100}")
        
        # 2. Sizes (Column 4)
        dv_sizes = DataValidation(type="list", formula1='"PP,P,M,G,GG"', allow_blank=True)
        dv_sizes.error = 'Escolha um porte da lista (PP, P, M, G ou GG)'
        dv_sizes.errorTitle = 'Porte Inválido'
        dv_sizes.prompt = 'Selecione o porte atendido'
        dv_sizes.promptTitle = 'Porte'
        ws.add_data_validation(dv_sizes)
        dv_sizes.add(f"D2:D{ws.max_row + 100}")

        # 3. Coat Types (Column 5)
        dv_coats = DataValidation(type="list", formula1='"Curta,Média,Longa,Dupla,Sem Pelo"', allow_blank=True)
        dv_coats.error = 'Escolha uma pelagem da lista'
        dv_coats.errorTitle = 'Pelagem Inválida'
        dv_coats.prompt = 'Selecione o tipo de pelagem (opcional)'
        dv_coats.promptTitle = 'Pelagem'
        ws.add_data_validation(dv_coats)
        dv_coats.add(f"E2:E{ws.max_row + 100}")

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
            
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    async def import_services_from_excel(self, db: Session, tenant_id: int, file_content: bytes, progress_callback=None):
        import openpyxl
        import io
        from app.modules.tenant_services.models import Service

        if not file_content:
            return {"created": 0, "updated": 0, "total": 0, "errors": ["Arquivo de planilha vazio"]}

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return {"created": 0, "updated": 0, "total": 0, "errors": [f"Erro ao ler arquivo Excel: {str(e)}"]}

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"created": 0, "updated": 0, "total": 0, "errors": ["A planilha está vazia"]}

        raw_headers = rows[0]
        headers = []
        for h in raw_headers:
            if h is None:
                headers.append("")
            else:
                h_str = str(h).strip().lower().replace("*", "").strip()
                if h_str in ("nome", "serviço", "nome_servico", "nome_do_servico"):
                    h_str = "servico"
                elif h_str in ("preço", "valor", "preco_venda"):
                    h_str = "preco"
                elif h_str in ("duração", "duracao", "tempo"):
                    h_str = "duracao_minutos"
                headers.append(h_str)

        required_cols = ["servico", "preco"]
        missing_cols = [col for col in required_cols if col not in headers]
        if missing_cols:
            return {
                "created": 0, "updated": 0, "total": 0,
                "errors": [f"Colunas obrigatórias ausentes na planilha: {', '.join(missing_cols)}"]
            }


        non_empty_rows = [
            rv for rv in rows[1:]
            if any(v is not None and str(v).strip() != "" for v in rv)
        ]
        total_rows = len(non_empty_rows)

        if progress_callback and total_rows > 0:
            progress_callback(0, total_rows, 0, 0)

        # Pre-fetch existing services (1 query)
        existing_services_q = db.query(Service).filter(Service.tenant_id == tenant_id).all()
        existing_services: dict[tuple, Service] = {
            (
                s.name.lower().strip(),
                (s.species or "").lower().strip(),
                (s.size or "").upper().strip(),
                (s.coat_type or "").lower().strip()
            ): s for s in existing_services_q
        }

        created_count = 0
        updated_count = 0
        processed_count = 0
        errors = []

        def parse_money(val):
            if val is None or str(val).strip() == "":
                return None
            if isinstance(val, (int, float)):
                return int(round(val * 100))
            try:
                cleaned = str(val).replace("R$", "").strip()
                if "," in cleaned and "." in cleaned:
                    if cleaned.rfind(",") > cleaned.rfind("."):
                        cleaned = cleaned.replace(".", "").replace(",", ".")
                    else:
                        cleaned = cleaned.replace(",", "")
                elif "," in cleaned:
                    cleaned = cleaned.replace(",", ".")
                return int(round(float(cleaned) * 100))
            except (ValueError, TypeError):
                return None

        def clean_str(val):
            if val is None or str(val).strip() == "":
                return None
            return str(val).strip()

        def parse_species(val):
            if val is None or str(val).strip() == "":
                return None
            cleaned = str(val).strip().lower()
            if "canino" in cleaned or "cão" in cleaned or "cao" in cleaned or "cachorro" in cleaned:
                return "Canino"
            if "felino" in cleaned or "gato" in cleaned:
                return "Felino"
            if "exotico" in cleaned or "exótico" in cleaned:
                return "Exoticos"
            return None

        def parse_size(val):
            if val is None or str(val).strip() == "":
                return None
            v = str(val).strip().lower()
            if v in ["pp", "mini", "micro"]:
                return "PP"
            elif v in ["p", "pequeno", "peq"]:
                return "P"
            elif v in ["m", "medio", "médio"]:
                return "M"
            elif v in ["g", "grande"]:
                return "G"
            elif v in ["gg", "gigante", "extra grande"]:
                return "GG"
            if "(pp)" in v or "mini" in v: return "PP"
            if "(p)" in v or "pequeno" in v: return "P"
            if "(m)" in v or "médio" in v or "medio" in v: return "M"
            if "(g)" in v or "grande" in v: return "G"
            if "(gg)" in v or "gigante" in v: return "GG"
            cleaned = str(val).strip().upper()
            if cleaned in ["PP", "P", "M", "G", "GG"]:
                return cleaned
            return None

        def parse_coat_type(val):
            if val is None or str(val).strip() == "":
                return None
            cleaned = str(val).strip().lower()
            coat_map = {
                "curta": "short", "média": "medium", "media": "medium",
                "longa": "long", "dupla": "double", "sem pelo": "hairless"
            }
            return coat_map.get(cleaned, None)

        CHUNK_SIZE = 200
        new_services: list[Service] = []

        def _flush_chunk():
            nonlocal new_services
            if new_services:
                db.add_all(new_services)
                db.flush()
                new_services = []
            db.commit()

        for row_idx, row_values in enumerate(non_empty_rows, start=2):
            try:
                row = {h: val for h, val in zip(headers, row_values) if h}

                servico = clean_str(row.get("servico"))
                if not servico:
                    raise ValueError("Nome do serviço é obrigatório")

                especie_val = clean_str(row.get("especie"))
                especie = parse_species(especie_val) if especie_val else None

                porte_val = clean_str(row.get("porte"))
                porte = parse_size(porte_val) if porte_val else None

                coat_val = clean_str(row.get("pelagem"))
                coat_type = parse_coat_type(coat_val)

                raw_price = row.get("preco")
                price_cents = parse_money(raw_price)
                if price_cents is None:
                    raise ValueError("Preço é obrigatório e deve ser um valor válido")

                raw_duration = row.get("duracao_minutos")
                duration_minutes = None
                if raw_duration is not None and str(raw_duration).strip() != "":
                    try:
                        duration_minutes = int(float(str(raw_duration)))
                    except Exception:
                        raise ValueError("Duração em minutos deve ser um número inteiro")

                desc = clean_str(row.get("descricao"))

                key = (
                    servico.lower().strip(),
                    (especie or "").lower().strip(),
                    (porte or "").upper().strip(),
                    (coat_type or "").lower().strip()
                )

                service_obj = existing_services.get(key)
                if not service_obj:
                    service_obj = Service(
                        tenant_id=tenant_id,
                        name=servico,
                        description=desc,
                        species=especie,
                        size=porte,
                        coat_type=coat_type,
                        price_cents=price_cents,
                        duration_minutes=duration_minutes,
                        is_active=True
                    )
                    new_services.append(service_obj)
                    existing_services[key] = service_obj
                    created_count += 1
                else:
                    _changed = False
                    for attr, val in [
                        ("price_cents", price_cents),
                        ("duration_minutes", duration_minutes),
                        ("description", desc),
                    ]:
                        if val is not None and getattr(service_obj, attr, None) != val:
                            setattr(service_obj, attr, val)
                            _changed = True
                    if _changed:
                        updated_count += 1

                processed_count += 1

            except Exception as e:
                errors.append(f"Linha {row_idx}: {str(e)}")

            if (row_idx - 1) % CHUNK_SIZE == 0:
                _flush_chunk()
                if progress_callback and total_rows > 0:
                    pct = min(int((row_idx - 1) / total_rows * 100), 99)
                    progress_callback(row_idx - 1, total_rows, processed_count, pct)

        _flush_chunk()

        return {
            "created": created_count,
            "updated": updated_count,
            "total": total_rows,
            "errors": errors,
        }

    async def import_services_from_excel_background(
        self,
        job_id: str,
        tenant_id: int,
        file_content: bytes,
    ) -> None:
        from app.modules.clients.import_jobs import update_job
        from app.config.database import SessionLocal

        db = SessionLocal()
        try:
            update_job(job_id, status="running", progress=0)

            def _on_progress(processed: int, total: int, count: int, pct: int):
                update_job(job_id, progress=pct, imported=count, total=total)

            result = await self.import_services_from_excel(
                db, tenant_id, file_content,
                progress_callback=_on_progress,
            )
            update_job(
                job_id,
                status="done",
                progress=100,
                created=result["created"],
                updated=result["updated"],
                total=result.get("total", 0),
                errors=result["errors"],
            )
        except Exception as e:
            update_job(job_id, status="error", errors=[str(e)])
        finally:
            db.close()

    def export_to_excel(self, db: Session, tenant_id: int) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Serviços"
        
        headers = [
            "ID", "Nome do Serviço", "Espécie", "Porte", "Pelagem", "Duração (min)", "Preço", "Descrição", "Ativo"
        ]
        ws.append(headers)
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        services = self.repo.list(db, tenant_id)
        
        coat_map = {
            "short": "Curta",
            "medium": "Média",
            "long": "Longa",
            "hairless": "Sem Pelo"
        }
        for s in services:
            species_val = s.species.value if hasattr(s.species, "value") else (s.species or "Todos")
            size_val = s.size.value if hasattr(s.size, "value") else (s.size or "Todos")
            coat_val = coat_map.get(s.coat_type, s.coat_type) if s.coat_type else "Todos"
            
            ws.append([
                s.id, s.name, species_val, size_val, coat_val,
                s.duration_minutes or "", float(s.price_cents) / 100 if s.price_cents else 0.0,
                s.description or "", "Sim" if s.is_active else "Não"
            ])
                
        out = BytesIO()
        wb.save(out)
        return out.getvalue()
