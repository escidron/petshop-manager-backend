import io
import re
import openpyxl
from io import BytesIO
from typing import Dict, List, Any, Optional
from datetime import date, datetime


# ---------------------------------------------------------------------------
# ZERO-AS-NULL SENTINEL — many legacy systems use "0" for empty fields
# ---------------------------------------------------------------------------
_ZERO_NULL_SENTINELS = {"0", "#n/a", "#na", "n/a", "na", "-", "--", "---", "none", "null", ""}


def _is_zero_null(val: Any) -> bool:
    """Returns True if the value is a legacy-system 'empty' sentinel (0, #N/A, etc.)."""
    if val is None:
        return True
    if isinstance(val, (int, float)) and val == 0:
        return True
    s = str(val).strip().lower()
    return s in _ZERO_NULL_SENTINELS


def normalize_phone(val: Any) -> Optional[str]:
    """
    Intelligently cleans and normalizes phone numbers.
    Handles +55, 5511999998888, (11) 99999-8888, 011999998888, 11999998888, etc.
    Returns format (XX) 9XXXX-XXXX or (XX) XXXX-XXXX.
    Discards legacy "0" empty-value sentinel.
    """
    if _is_zero_null(val):
        return None

    raw = str(val).strip()
    digits = "".join(filter(str.isdigit, raw))

    if not digits or digits == "0":
        return None

    # Strip country code 55 if present
    if digits.startswith("55") and len(digits) in (12, 13):
        digits = digits[2:]

    # Strip leading zero if present (e.g., 011999998888)
    if digits.startswith("0") and len(digits) in (11, 12):
        digits = digits[1:]

    # Standardize length: 10 digits (landline) or 11 digits (mobile)
    if len(digits) == 11:
        return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"
    elif len(digits) == 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"

    # Fallback to raw if we can't normalize
    return raw


def clean_str(val: Any) -> Optional[str]:
    """
    Cleans a generic string field.
    - Discards legacy "0" / "#N/A" null sentinels
    - Strips leading/trailing whitespace and embedded newlines
    - Converts float integers (1.0) to "1"
    """
    if _is_zero_null(val):
        return None

    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).strip()

    val_str = str(val).strip()
    # Remove embedded newlines/carriage-returns (corrupted multi-line cells)
    val_str = val_str.replace("\r\n", " ").replace("\n", " ").replace("\r", " ").strip()

    if not val_str or val_str.lower() in _ZERO_NULL_SENTINELS:
        return None

    return val_str


def clean_email(val: Any) -> Optional[str]:
    """
    Returns a sanitized email or None.
    Discards: "0", "#N/A", strings without "@", multi-line garbage.
    """
    s = clean_str(val)
    if not s:
        return None
    # Must contain @ and at least one dot after it
    if "@" not in s:
        return None
    local, _, domain = s.partition("@")
    if not local or "." not in domain:
        return None
    # Discard obviously bad patterns like "email\n0" artifacts
    if re.search(r"[\n\r\t]", s):
        return None
    return s.lower().strip()


def clean_cep(val: Any) -> Optional[str]:
    """
    Returns sanitized CEP (8 digits, adds leading zero if 7 digits).
    Discards: "0", "00000000", non-numeric-looking strings.
    """
    if _is_zero_null(val):
        return None
    digits = "".join(filter(str.isdigit, str(val)))
    if not digits or all(c == "0" for c in digits):
        return None
    # Fix 7-digit CEPs missing leading zero
    if len(digits) == 7:
        digits = "0" + digits
    if len(digits) != 8:
        return None
    return digits


def clean_address_number(val: Any) -> Optional[str]:
    """
    Cleans street number field.
    - Discards "0" sentinel
    - If the value looks like a full address (contains street keywords), returns None
      (it was probably a copy-paste bug in the source system)
    """
    if _is_zero_null(val):
        return None
    s = str(val).strip()
    # Discard if it contains address-like keywords (source system bug: full addr in number field)
    address_keywords = re.compile(
        r"\b(rua|av|avenida|alameda|travessa|estrada|rodovia|praca|praça|r\.)\b",
        re.IGNORECASE,
    )
    if address_keywords.search(s):
        return None
    # If it has "Apto"/"Ap" in it, split and take the numeric part only
    apto_match = re.match(r"^(\d+)\s*(apto?|ap\.?|apartamento)\s*(.*)$", s, re.IGNORECASE)
    if apto_match:
        return apto_match.group(1).strip()
    return s[:20] or None  # cap at 20 chars


def clean_address_str(val: Any) -> Optional[str]:
    """
    Cleans a generic address text field (logradouro, bairro, cidade).
    Discards "0", purely numeric zero-like values.
    """
    if _is_zero_null(val):
        return None
    s = str(val).strip()
    # Discard if it's just zeros, dashes or "0 - SP" type patterns
    if re.fullmatch(r"[0\s\-\/]+", s):
        return None
    if s.lower() in _ZERO_NULL_SENTINELS:
        return None
    # Remove embedded newlines
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ").strip()
    return s or None


def clean_state(val: Any) -> Optional[str]:
    """Returns a 2-letter UF or None. Discards "0", "SP" preceded by garbage, etc."""
    if _is_zero_null(val):
        return None
    s = str(val).strip().upper()
    # Extract 2-letter state code from patterns like "0 - SP" or "SP"
    uf_match = re.search(r"\b([A-Z]{2})\b", s)
    if uf_match:
        uf = uf_match.group(1)
        # Validate it's a real Brazilian state
        valid_ufs = {
            "AC", "AL", "AM", "AP", "BA", "CE", "DF", "ES", "GO",
            "MA", "MG", "MS", "MT", "PA", "PB", "PE", "PI", "PR",
            "RJ", "RN", "RO", "RR", "RS", "SC", "SE", "SP", "TO",
        }
        if uf in valid_ufs:
            return uf
    return None


def normalize_porte(val: Any) -> Optional[str]:
    if _is_zero_null(val):
        return None
    cleaned = str(val).strip().lower()
    if "mini" in cleaned or "pp" in cleaned or "micro" in cleaned:
        return "PP"
    elif "pequeno" in cleaned or " p " in f" {cleaned} " or cleaned == "p":
        return "P"
    elif "médio" in cleaned or "medio" in cleaned or " m " in f" {cleaned} " or cleaned == "m":
        return "M"
    elif "grande" in cleaned or " g " in f" {cleaned} " or cleaned == "g":
        return "G"
    elif "gigante" in cleaned or "gg" in cleaned or "extra" in cleaned:
        return "GG"
    return str(val).strip()[:5]


def parse_species(val: Any) -> Optional[str]:
    """
    Normalizes species from any legacy system format to our standard values.
    Handles: "Canideo", "canídeo", "Cão", "cão", "Cachorro", "Dog", "Canino" → "Canino"
             "Felino", "Gato", "Cat" → "Felino"
             "Exotico", "Exótico", "Ave", "Peixe", "Réptil", "Jabuti" → "Exoticos"
    """
    if _is_zero_null(val):
        return None
    cleaned = str(val).strip().lower()
    # Canine variants — including the "canideo/canídeo" typo seen in real data
    if any(k in cleaned for k in [
        "canino", "canídeo", "canideo", "cão", "cao", "cachorro", "dog",
        "vira-lata", "viralata", "can ", "canil"
    ]):
        return "Canino"
    # Feline variants
    if any(k in cleaned for k in ["felino", "felídeo", "felideo", "gato", "cat", "felidae", "gatinho", "miau"]):
        return "Felino"
    # Exotic — broad catch
    if any(k in cleaned for k in [
        "exotico", "exótico", "exot", "ave", "peixe", "reptil", "réptil",
        "jabuti", "tartaruga", "coelho", "hamster", "porquinho", "guinea",
        "papagaio", "calopsita", "periquito", "arara", "cobra", "lagarto",
        "iguana", "furão", "furao",
    ]):
        return "Exoticos"
    # Fallback: return raw value capitalised (don't lose data)
    return str(val).strip()


def parse_gender(val: Any) -> str:
    """
    Normalizes gender to our internal values: "male" | "female" | "unknown".
    Handles: "M", "F", "Macho", "Fêmea", "Femea", "male", "female", "0/1" etc.
    """
    if _is_zero_null(val):
        return "unknown"
    cleaned = str(val).strip().lower()
    if cleaned in ("m", "macho", "male", "masculino", "1"):
        return "male"
    if cleaned in ("f", "fêmea", "femea", "female", "feminino", "2"):
        return "female"
    if "macho" in cleaned:
        return "male"
    if "fêmea" in cleaned or "femea" in cleaned:
        return "female"
    return "unknown"


def parse_deceased(val: Any) -> bool:
    """
    Normalizes death status from legacy switch/boolean/text values.
    Truthy: 1, "1", "sim", "s", "yes", "true", "falecido", "morto"
    Falsy (default): 0, "0", "não", "nao", "n", "no", "false", None
    """
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return int(val) == 1
    cleaned = str(val).strip().lower()
    return cleaned in ("1", "sim", "s", "yes", "true", "falecido", "morto", "obito", "óbito")


def parse_neutered(val: Any) -> Optional[bool]:
    """Returns True/False/None for castration status."""
    if _is_zero_null(val):
        return None
    cleaned = str(val).strip().lower()
    if cleaned in ("sim", "s", "yes", "1", "true", "castrado", "castrada"):
        return True
    if cleaned in ("não", "nao", "n", "no", "0", "false", "inteiro", "inteira"):
        return False
    return None


def parse_coat_type(val: Any) -> Optional[str]:
    if _is_zero_null(val):
        return None
    cleaned = str(val).strip().lower()
    if "curta" in cleaned or "curto" in cleaned or "short" in cleaned:
        return "short"
    if "média" in cleaned or "media" in cleaned or "medium" in cleaned:
        return "medium"
    if "longa" in cleaned or "longo" in cleaned or "long" in cleaned:
        return "long"
    if "dupla" in cleaned or "double" in cleaned:
        return "double"
    if "sem pelo" in cleaned or "hairless" in cleaned or "calvo" in cleaned:
        return "hairless"
    return clean_str(val)


def parse_age_unit(val: Any) -> Optional[str]:
    if _is_zero_null(val):
        return None
    cleaned = str(val).strip().lower()
    if any(k in cleaned for k in ("meses", "mês", "mes", "month")):
        return "months"
    if any(k in cleaned for k in ("anos", "ano", "year")):
        return "years"
    return None


class MigrationService:
    def inspect_excel_headers(self, file_content: bytes) -> List[str]:
        """
        Reads row 1 headers of an uploaded Excel file.
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        raw_headers = rows[0]
        headers = []
        for h in raw_headers:
            if h is not None and str(h).strip() != "":
                headers.append(str(h).strip())
        return headers

    def parse_excel_rows(self, file_content: bytes) -> List[Dict[str, Any]]:
        """
        Parses an Excel file into a list of row dicts keyed by header name.
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return []

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]

        parsed = []
        for row_values in rows[1:]:
            row_dict = {}
            has_val = False
            for h, val in zip(headers, row_values):
                if h:
                    row_dict[h] = val
                    if val is not None and str(val).strip() not in ("", "0"):
                        has_val = True
            if has_val:
                parsed.append(row_dict)
        return parsed

    def _apply_field_transform(self, target_field: str, val: Any) -> Any:
        """
        Applies the correct sanitization/normalization for each target field type.
        This is the single source-of-truth for how each field is cleaned.
        """
        if target_field in ("cliente_telefone", "cliente_telefone_2", "cliente_telefone_3"):
            return normalize_phone(val)
        if target_field == "cliente_email":
            return clean_email(val)
        if target_field == "cliente_cep":
            return clean_cep(val)
        if target_field == "cliente_numero":
            return clean_address_number(val)
        if target_field in ("cliente_logradouro", "cliente_bairro", "cliente_cidade", "cliente_complemento"):
            return clean_address_str(val)
        if target_field == "cliente_estado":
            return clean_state(val)
        if target_field == "pet_porte":
            return normalize_porte(val)
        if target_field == "pet_especie":
            return parse_species(val)
        if target_field == "pet_genero":
            return parse_gender(val)
        if target_field == "pet_castrado":
            return parse_neutered(val)
        if target_field == "pet_obito":
            return parse_deceased(val)
        if target_field == "pet_tipo_pelagem":
            return parse_coat_type(val)
        if target_field == "pet_unidade_idade":
            return parse_age_unit(val)
        # Generic string clean for everything else
        return clean_str(val)

    def map_single_file(
        self,
        file_content: bytes,
        column_mapping: Dict[str, str],  # target_field -> source_header
        default_values: Dict[str, str] = None,  # target_field -> default_str
    ) -> List[Dict[str, Any]]:
        """
        Transforms a single competitor Excel file into system-standard rows.
        """
        default_values = default_values or {}
        source_rows = self.parse_excel_rows(file_content)

        transformed = []
        for row in source_rows:
            item = {}
            for target_field, source_col in column_mapping.items():
                if source_col and source_col in row:
                    val = row[source_col]
                    item[target_field] = self._apply_field_transform(target_field, val)
                else:
                    item[target_field] = default_values.get(target_field)

            # Apply defaults for empty mapped fields if default exists
            for k, default_val in default_values.items():
                if not item.get(k) and default_val:
                    item[k] = default_val

            if item.get("cliente_nome") or item.get("pet_nome"):
                transformed.append(item)

        return transformed

    def map_two_files(
        self,
        clients_file_content: bytes,
        pets_file_content: bytes,
        client_link_key: str,  # e.g., 'ID Cliente' in clients file
        pet_link_key: str,     # e.g., 'ID Dono' in pets file
        client_column_mapping: Dict[str, str],  # cliente_* target -> client file header
        pet_column_mapping: Dict[str, str],     # pet_* target -> pet file header
        default_values: Dict[str, str] = None,
    ) -> List[Dict[str, Any]]:
        """
        Joins separate Clients and Pets Excel files on matching ID key into system-standard rows.
        """
        default_values = default_values or {}
        client_rows = self.parse_excel_rows(clients_file_content)
        pet_rows = self.parse_excel_rows(pets_file_content)

        # Index client rows by link key
        client_by_key = {}
        for cr in client_rows:
            raw_key = cr.get(client_link_key)
            if raw_key is not None and str(raw_key).strip() != "":
                client_by_key[str(raw_key).strip()] = cr

        transformed = []
        for pr in pet_rows:
            raw_link = pr.get(pet_link_key)
            link_str = str(raw_link).strip() if raw_link is not None else ""
            client_row = client_by_key.get(link_str, {})

            item = {}
            # Map client fields
            for target_field, source_col in client_column_mapping.items():
                if source_col and source_col in client_row:
                    val = client_row[source_col]
                    item[target_field] = self._apply_field_transform(target_field, val)
                else:
                    item[target_field] = default_values.get(target_field)

            # Map pet fields
            for target_field, source_col in pet_column_mapping.items():
                if source_col and source_col in pr:
                    val = pr[source_col]
                    item[target_field] = self._apply_field_transform(target_field, val)
                else:
                    item[target_field] = default_values.get(target_field)

            # Apply defaults
            for k, default_val in default_values.items():
                if not item.get(k) and default_val:
                    item[k] = default_val

            if item.get("cliente_nome") or item.get("pet_nome"):
                transformed.append(item)

        return transformed

    def generate_converted_excel(self, items: List[Dict[str, Any]]) -> bytes:
        """
        Generates an Excel workbook formatted in our system's standard template schema from mapped items.
        Includes all fields supported by the Client and Pet models.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes e Pets"

        headers = [
            # Client fields
            "cliente_nome *", "cliente_telefone *", "cliente_email", "cliente_documento",
            "cliente_data_nascimento",
            "cliente_telefone_2", "cliente_nome_contato_2",
            "cliente_telefone_3", "cliente_nome_contato_3",
            "cliente_instagram", "cliente_facebook",
            "cliente_cep", "cliente_logradouro", "cliente_numero",
            "cliente_complemento", "cliente_bairro", "cliente_cidade", "cliente_estado",
            # Pet fields
            "pet_nome *", "pet_especie *", "pet_raca", "pet_tipo_pelagem", "pet_cor_pelagem",
            "pet_genero", "pet_porte", "pet_castrado", "pet_obito",
            "pet_data_nascimento", "pet_idade_aproximada", "pet_unidade_idade", "pet_observacoes",
        ]
        ws.append(headers)

        header_keys = [
            "cliente_nome", "cliente_telefone", "cliente_email", "cliente_documento",
            "cliente_data_nascimento",
            "cliente_telefone_2", "cliente_nome_contato_2",
            "cliente_telefone_3", "cliente_nome_contato_3",
            "cliente_instagram", "cliente_facebook",
            "cliente_cep", "cliente_logradouro", "cliente_numero",
            "cliente_complemento", "cliente_bairro", "cliente_cidade", "cliente_estado",
            "pet_nome", "pet_especie", "pet_raca", "pet_tipo_pelagem", "pet_cor_pelagem",
            "pet_genero", "pet_porte", "pet_castrado", "pet_obito",
            "pet_data_nascimento", "pet_idade_aproximada", "pet_unidade_idade", "pet_observacoes",
        ]

        for item in items:
            row = [item.get(k, "") or "" for k in header_keys]
            ws.append(row)

        out = BytesIO()
        wb.save(out)
        return out.getvalue()
