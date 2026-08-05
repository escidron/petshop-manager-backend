import csv
import io
from fastapi import HTTPException, status
from sqlalchemy.orm import Session
from .repository import ProductRepository
from .inventory_repository import InventoryRepository
from .schemas import ProductCreate, ProductUpdate


class ProductService:
    def __init__(self):
        self.repository = ProductRepository()
        self.inventory_repository = InventoryRepository()

    def create_product(self, db: Session, tenant_id: int, data: ProductCreate):
        return self.repository.create(db, tenant_id, data)

    def get_product(self, db: Session, tenant_id: int, product_id: int):
        product = self.repository.get_by_id(db, tenant_id, product_id)
        if not product:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Produto não encontrado",
            )
        return product

    def list_products(self, db: Session, tenant_id: int, exclude_internal: bool = False):
        return self.repository.list(db, tenant_id, exclude_internal=exclude_internal)

    def update_product(self, db: Session, tenant_id: int, product_id: int, data: ProductUpdate):
        product = self.get_product(db, tenant_id, product_id)
        return self.repository.update(db, product, data)

    def delete_product(self, db: Session, tenant_id: int, product_id: int):
        product = self.get_product(db, tenant_id, product_id)
        self.repository.delete(db, product)

    def adjust_stock(
        self, 
        db: Session, 
        tenant_id: int, 
        product_id: int, 
        quantity_change: int, 
        change_type: str, 
        notes: str | None = None
    ):
        product = self.get_product(db, tenant_id, product_id)
        product.quantity += quantity_change
        
        self.inventory_repository.create_log(
            db, tenant_id, product_id, quantity_change, change_type, notes
        )
        return product

    async def import_products_from_excel(self, db: Session, tenant_id: int, file_content: bytes, progress_callback=None):

        import openpyxl
        import io
        from app.modules.products.models import Product

        if not file_content:
            return {"created": 0, "updated": 0, "total": 0, "errors": ["Arquivo de planilha vazio"]}

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return {"created": 0, "updated": 0, "total": 0, "errors": [f"Erro ao ler arquivo Excel: {str(e)}"]}

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"created": 0, "updated": 0, "total": 0, "errors": ["A planilha está vazia"]}

        raw_headers = rows[0]
        headers = []
        for h in raw_headers:
            if h is None:
                headers.append("")
            else:
                headers.append(str(h).strip().lower().replace("*", "").strip())

        required_cols = ["nome", "preco_venda"]
        missing_cols = [col for col in required_cols if col not in headers]
        if missing_cols:
            return {
                "created": 0, "updated": 0, "total": 0,
                "errors": [f"Colunas obrigatórias ausentes na planilha: {', '.join(missing_cols)}"]
            }

        non_empty_rows = [
            rv for rv in rows[1:]
            if any(v is not None and str(v).strip() != "" for v in rv)
        ]
        total_rows = len(non_empty_rows)

        if progress_callback and total_rows > 0:
            progress_callback(0, total_rows, 0, 0)

        # Pre-fetch existing products (1 query)
        existing_prods = db.query(Product).filter(Product.tenant_id == tenant_id).all()
        existing_by_barcode: dict[str, Product] = {p.barcode: p for p in existing_prods if p.barcode}
        existing_by_sku: dict[str, Product] = {p.sku: p for p in existing_prods if p.sku}
        existing_by_name: dict[str, Product] = {p.name.lower().strip(): p for p in existing_prods if p.name}

        created_count = 0
        updated_count = 0
        processed_count = 0
        errors = []

        def parse_money(val):
            if val is None or str(val).strip() == "":
                return None
            if isinstance(val, (int, float)):
                return int(round(val * 100))
            try:
                cleaned = str(val).replace("R$", "").strip()
                if "," in cleaned and "." in cleaned:
                    if cleaned.rfind(",") > cleaned.rfind("."):
                        cleaned = cleaned.replace(".", "").replace(",", ".")
                    else:
                        cleaned = cleaned.replace(",", "")
                elif "," in cleaned:
                    cleaned = cleaned.replace(",", ".")
                return int(round(float(cleaned) * 100))
            except (ValueError, TypeError):
                return None

        def clean_str(val):
            if val is None or str(val).strip() == "":
                return None
            if isinstance(val, float) and val.is_integer():
                return str(int(val))
            return str(val).strip()

        def parse_unit(val):
            if val is None or str(val).strip() == "":
                return "UN"
            cleaned = str(val).strip().lower()
            unit_map = {
                "unidade (un)": "UN", "grama (g)": "g", "quilograma (kg)": "kg",
                "mililitro (ml)": "ml", "litro (l)": "L", "pacote (paq)": "PAQ",
                "caixa (cx)": "CX", "un": "UN", "g": "g", "kg": "kg",
                "ml": "ml", "l": "L", "paq": "PAQ", "cx": "CX"
            }
            return unit_map.get(cleaned, "UN")

        CHUNK_SIZE = 200
        new_products: list[Product] = []

        def _flush_chunk():
            nonlocal new_products
            if new_products:
                db.add_all(new_products)
                db.flush()
                new_products = []
            db.commit()

        for row_idx, row_values in enumerate(non_empty_rows, start=2):
            try:
                row = {h: val for h, val in zip(headers, row_values) if h}

                name = clean_str(row.get("nome"))
                if not name:
                    raise ValueError("Nome do produto é obrigatório")

                raw_price = row.get("preco_venda")
                price = parse_money(raw_price)
                if price is None:
                    raise ValueError("Preço de venda é obrigatório e deve ser um valor válido")

                cost = parse_money(row.get("custo"))
                barcode = clean_str(row.get("codigo_barras"))
                sku = clean_str(row.get("sku"))
                ncm = clean_str(row.get("ncm"))
                description = clean_str(row.get("descricao"))
                category = clean_str(row.get("categoria"))
                unit = parse_unit(row.get("unidade"))

                try:
                    quantity = int(row.get("quantidade") or 0)
                except (ValueError, TypeError):
                    quantity = 0

                try:
                    min_stock = int(row.get("estoque_minimo") or 0)
                except (ValueError, TypeError):
                    min_stock = 0

                # Match existing product
                product = None
                if barcode and barcode in existing_by_barcode:
                    product = existing_by_barcode[barcode]
                elif sku and sku in existing_by_sku:
                    product = existing_by_sku[sku]
                elif name.lower() in existing_by_name:
                    product = existing_by_name[name.lower()]

                if not product:
                    product = Product(
                        tenant_id=tenant_id,
                        name=name,
                        price=price,
                        cost=cost,
                        barcode=barcode,
                        sku=sku,
                        ncm=ncm,
                        description=description,
                        category=category,
                        unit=unit,
                        quantity=quantity,
                        min_stock=min_stock,
                        cest=clean_str(row.get("cest")),
                        cfop=clean_str(row.get("cfop")),
                        csosn=clean_str(row.get("csosn")),
                        cst_pis=clean_str(row.get("cst_pis")),
                        cst_cofins=clean_str(row.get("cst_cofins")),
                        is_active=True,
                    )
                    new_products.append(product)
                    if barcode:
                        existing_by_barcode[barcode] = product
                    if sku:
                        existing_by_sku[sku] = product
                    existing_by_name[name.lower()] = product
                    created_count += 1
                else:
                    _changed = False
                    for attr, val in [
                        ("price", price), ("cost", cost), ("sku", sku),
                        ("ncm", ncm), ("description", description),
                        ("category", category), ("unit", unit),
                        ("quantity", quantity), ("min_stock", min_stock),
                    ]:
                        if val is not None and getattr(product, attr, None) != val:
                            setattr(product, attr, val)
                            _changed = True
                    if _changed:
                        updated_count += 1

                processed_count += 1

            except Exception as e:
                errors.append(f"Linha {row_idx}: {str(e)}")

            if (row_idx - 1) % CHUNK_SIZE == 0:
                _flush_chunk()
                if progress_callback and total_rows > 0:
                    pct = min(int((row_idx - 1) / total_rows * 100), 99)
                    progress_callback(row_idx - 1, total_rows, processed_count, pct)

        _flush_chunk()

        return {
            "created": created_count,
            "updated": updated_count,
            "total": total_rows,
            "errors": errors,
        }

    async def import_products_from_excel_background(
        self,
        job_id: str,
        tenant_id: int,
        file_content: bytes,
    ) -> None:
        from app.modules.clients.import_jobs import update_job
        from app.config.database import SessionLocal

        db = SessionLocal()
        try:
            update_job(job_id, status="running", progress=0)

            def _on_progress(processed: int, total: int, count: int, pct: int):
                update_job(job_id, progress=pct, imported=count, total=total)

            result = await self.import_products_from_excel(
                db, tenant_id, file_content,
                progress_callback=_on_progress,
            )
            update_job(
                job_id,
                status="done",
                progress=100,
                created=result["created"],
                updated=result["updated"],
                total=result.get("total", 0),
                errors=result["errors"],
            )
        except Exception as e:
            update_job(job_id, status="error", errors=[str(e)])
        finally:
            db.close()

    def generate_import_template_excel(self) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produtos"
        
        # Add headers
        headers = [
            "nome *", "preco_venda *", "unidade *", "quantidade *", "estoque_minimo *", "codigo_barras *", "ncm *",
            "sku", "descricao", "categoria", "custo", 
            "cest", "cfop", "csosn", "cst_pis", "cst_cofins"
        ]
        ws.append(headers)
        
        # Style headers (Segoe UI, Blue color theme)
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        required_header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3")
        )
        
        # Format header row
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            header_name = headers[col_idx - 1]
            if "*" in header_name:
                cell.fill = required_header_fill
            else:
                cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # Add example row
        example = [
            "Ração Premium Gato Adulto 1kg",
            89.90,
            "Unidade (UN)",
            20,
            5,
            "7891234567890",
            "3801.10.00",
            "RAC-PREM-GAT-1",
            "Ração premium para gatos adultos sabor salmão",
            "Ração",
            45.00,
            "01.001.00",
            "5102",
            "102",
            "01",
            "01"
        ]
        ws.append(example)
        
        # Format example row (row 2)
        ws.row_dimensions[2].height = 20
        for col_idx in range(1, len(example) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            
            # Alignments & formatting based on content
            if col_idx in [1, 4, 5]: # Text
                cell.alignment = left_align
            else: # Numbers / Codes / Selects
                cell.alignment = center_align
                
            # Number formats
            if col_idx in [2, 7]: # currency
                cell.number_format = "#,##0.00"
            elif col_idx in [8, 9]: # integer
                cell.number_format = "#,##0"
                
        # Data Validations (Dropdowns)
        # 1. Units (Column C - col_idx 3)
        dv_unit = DataValidation(type="list", formula1='"Unidade (UN),Grama (g),Quilograma (kg),Mililitro (ml),Litro (L),Pacote (PAQ),Caixa (CX)"', allow_blank=True)
        dv_unit.error = 'Escolha uma unidade da lista'
        dv_unit.errorTitle = 'Unidade Inválida'
        dv_unit.prompt = 'Selecione a unidade de medida do produto'
        dv_unit.promptTitle = 'Unidade de Medida'
        ws.add_data_validation(dv_unit)
        dv_unit.add("C2:C1000")
        
        # 2. Categories (Column F - col_idx 6)
        categories_list = "Ração,Acessórios,Higiene,Brinquedos,Medicamentos,Petiscos,Camas e Casinhas,Roupas,Aquarismo,Aves,Outros"
        dv_category = DataValidation(type="list", formula1=f'"{categories_list}"', allow_blank=True)
        dv_category.error = 'Escolha uma categoria da lista'
        dv_category.errorTitle = 'Categoria Inválida'
        dv_category.prompt = 'Selecione a categoria do produto'
        dv_category.promptTitle = 'Categoria'
        ws.add_data_validation(dv_category)
        dv_category.add("F2:F1000")
                
        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
            
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def export_to_excel(self, db: Session, tenant_id: int) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produtos"
        
        headers = [
            "ID", "Nome", "Código de Barras", "NCM", "SKU", "Categoria", "Unidade",
            "Preço Venda", "Preço Custo", "Estoque Atual", "Estoque Mínimo", "CEST", "CFOP", "CSOSN", "CST PIS", "CST COFINS", "Descrição", "Ativo"
        ]
        ws.append(headers)
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        products = self.repository.list(db, tenant_id)
        
        for p in products:
            ws.append([
                p.id, p.name, p.barcode or "", p.ncm or "", p.sku or "", p.category or "", p.unit or "",
                float(p.price), float(p.cost) if p.cost is not None else "", p.quantity, p.min_stock, 
                p.cest or "", p.cfop or "", p.csosn or "", p.cst_pis or "", p.cst_cofins or "", p.description or "",
                "Sim" if p.is_active else "Não"
            ])
                
        out = BytesIO()
        wb.save(out)
        return out.getvalue()
