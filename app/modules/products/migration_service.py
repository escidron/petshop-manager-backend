import io
import openpyxl
from io import BytesIO
from typing import Dict, List, Any, Optional

def clean_str(val: Any) -> Optional[str]:
    if val is None or str(val).strip() == "":
        return None
    val_str = str(val).strip()
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return val_str

def parse_num(val: Any) -> Optional[float]:
    if val is None or str(val).strip() == "":
        return None
    try:
        val_str = str(val).strip().replace("R$", "").replace("r$", "").replace(" ", "")
        if "," in val_str and "." in val_str:
            val_str = val_str.replace(".", "").replace(",", ".")
        elif "," in val_str:
            val_str = val_str.replace(",", ".")
        return float(val_str)
    except Exception:
        return None

class ProductMigrationService:
    def inspect_excel_headers(self, file_content: bytes) -> List[str]:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        return [str(h).strip() for h in rows[0] if h is not None and str(h).strip() != ""]

    def parse_excel_rows(self, file_content: bytes) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        parsed = []
        for r_vals in rows[1:]:
            r_dict = {}
            has_val = False
            for h, v in zip(headers, r_vals):
                if h:
                    r_dict[h] = v
                    if v is not None and str(v).strip() != "":
                        has_val = True
            if has_val:
                parsed.append(r_dict)
        return parsed

    def map_file(
        self,
        file_content: bytes,
        column_mapping: Dict[str, str],
        default_values: Dict[str, str] = None,
    ) -> List[Dict[str, Any]]:
        default_values = default_values or {}
        rows = self.parse_excel_rows(file_content)
        transformed = []
        for r in rows:
            item = {}
            for target_field, source_col in column_mapping.items():
                if source_col and source_col in r:
                    val = r[source_col]
                    if target_field in ("preco_custo", "preco_venda", "estoque_atual", "estoque_minimo"):
                        val = parse_num(val)
                    else:
                        val = clean_str(val)
                    item[target_field] = val
                else:
                    item[target_field] = default_values.get(target_field)

            for k, def_val in default_values.items():
                if not item.get(k) and def_val:
                    item[k] = def_val

            if item.get("nome"):
                transformed.append(item)
        return transformed

    def generate_converted_excel(self, items: List[Dict[str, Any]]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produtos"

        headers = [
            "nome *", "preco_venda *", "unidade *", "quantidade *", "estoque_minimo *", "codigo_barras *", "ncm *",
            "sku", "descricao", "categoria", "custo", 
            "cest", "cfop", "csosn", "cst_pis", "cst_cofins"
        ]
        ws.append(headers)

        # Mapping our internal item keys to the corresponding header column
        key_mapping = {
            "nome *": "nome",
            "preco_venda *": "preco_venda",
            "unidade *": "unidade",
            "quantidade *": "estoque_atual",
            "estoque_minimo *": "estoque_minimo",
            "codigo_barras *": "codigo_barras",
            "ncm *": "ncm",
            "sku": "sku",
            "descricao": "descricao",
            "categoria": "categoria",
            "custo": "preco_custo",
            "cest": "cest",
            "cfop": "cfop",
            "csosn": "csosn",
            "cst_pis": "cst_pis",
            "cst_cofins": "cst_cofins"
        }

        for item in items:
            row_data = []
            for h in headers:
                dict_key = key_mapping[h]
                row_data.append(item.get(dict_key, "") or "")
            ws.append(row_data)

        out = BytesIO()
        wb.save(out)
        return out.getvalue()
