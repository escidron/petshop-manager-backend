import io
from typing import Optional, List, Dict
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.modules.financial.models import DREAccount, DREEntry
from app.modules.financial.schemas import (
    DREAccountCreate,
    DREAccountUpdate,
    DREAccountResponse,
    DREEntryUpsert,
    DREEntryReplicate,
    DRERowData,
    DREGroupData,
    DRESummary,
    DREReportResponse,
)
from app.modules.financial.repository import FinancialRepository


class FinancialService:
    def __init__(self):
        self.repo = FinancialRepository()

    def get_dre_report(self, db: Session, tenant_id: int, year: int) -> DREReportResponse:
        # 1. Garante que as contas padrão existam
        accounts = self.repo.seed_default_accounts_if_needed(db, tenant_id)

        # 2. Busca dados automáticos do sistema
        sales_data = self.repo.get_sales_aggregated_by_month(db, tenant_id, year)
        cmv_data = self.repo.get_cmv_aggregated_by_month(db, tenant_id, year)
        commissions_data = self.repo.get_commissions_aggregated_by_month(db, tenant_id, year)

        # 3. Busca lançamentos manuais
        manual_entries = self.repo.get_entries_for_year(db, tenant_id, year)
        entries_map: Dict[int, Dict[int, float]] = {}
        for e in manual_entries:
            if e.account_id not in entries_map:
                entries_map[e.account_id] = {}
            entries_map[e.account_id][e.competence_month] = float(e.amount or 0.0)

        # 4. Agrupa contas por grupo
        groups_config = [
            ("gross_revenue", "(+) RECEITA BRUTA DE VENDAS", "TOTAL RECEITA BRUTA"),
            ("cmv", "(-) CUSTO MERCADORIA E SERVIÇO VENDIDO (CMV / CSP)", "TOTAL CMV / CSP"),
            ("fixed_expense", "(-) DESPESAS OPERACIONAIS FIXAS", "TOTAL DESPESAS FIXAS"),
            ("variable_expense", "(-) DESPESAS OPERACIONAIS VARIÁVEIS", "TOTAL DESPESAS VARIÁVEIS"),
            ("financial_result", "(+/-) RESULTADOS NÃO OPERACIONAIS / FINANCEIROS", "TOTAL RESULTADO FINANCEIRO"),
        ]

        accounts_by_group: Dict[str, List[DREAccount]] = {g[0]: [] for g in groups_config}
        for acc in accounts:
            if acc.is_active and acc.group_type in accounts_by_group:
                accounts_by_group[acc.group_type].append(acc)

        # 5. Constrói linhas de cada conta
        group_rows_map: Dict[str, List[DRERowData]] = {g[0]: [] for g in groups_config}
        group_monthly_totals: Dict[str, Dict[int, float]] = {
            g[0]: {m: 0.0 for m in range(1, 13)} for g in groups_config
        }

        for group_type, title, subtotal_name in groups_config:
            for acc in accounts_by_group[group_type]:
                monthly_amounts: Dict[int, float] = {}

                for m in range(1, 13):
                    # Se tiver lançamento manual explícito, tem prioridade
                    if acc.id in entries_map and m in entries_map[acc.id]:
                        val = entries_map[acc.id][m]
                    elif acc.is_system and acc.system_source:
                        if acc.system_source == "sales_products":
                            val = sales_data["sales_products"].get(m, 0.0)
                        elif acc.system_source == "sales_services":
                            val = sales_data["sales_services"].get(m, 0.0)
                        elif acc.system_source == "cmv_products":
                            val = cmv_data.get(m, 0.0)
                        elif acc.system_source == "commissions":
                            val = commissions_data.get(m, 0.0)
                        else:
                            val = 0.0
                    else:
                        val = 0.0

                    monthly_amounts[m] = round(val, 2)
                    group_monthly_totals[group_type][m] += val

                total_amount = round(sum(monthly_amounts.values()), 2)
                monthly_avg = round(total_amount / 12.0, 2)

                row = DRERowData(
                    id=f"acc-{acc.id}",
                    account_id=acc.id,
                    name=acc.name,
                    code=acc.code,
                    group_type=group_type,
                    is_system=acc.is_system,
                    system_source=acc.system_source,
                    is_header=False,
                    is_subtotal=False,
                    is_result=False,
                    is_percentage_row=False,
                    is_editable=True,
                    display_order=acc.order_index,
                    monthly_amounts=monthly_amounts,
                    monthly_percentages={},  # preenchido no passo da análise vertical
                    total_amount=total_amount,
                    total_percentage=0.0,
                    monthly_average=monthly_avg,
                )
                group_rows_map[group_type].append(row)

        # 6. Totais e Cálculos Gerenciais
        gross_rev_m = group_monthly_totals["gross_revenue"]
        cmv_m = group_monthly_totals["cmv"]
        fixed_exp_m = group_monthly_totals["fixed_expense"]
        var_exp_m = group_monthly_totals["variable_expense"]
        fin_res_m = group_monthly_totals["financial_result"]

        gross_margin_m = {m: round(gross_rev_m[m] - cmv_m[m], 2) for m in range(1, 13)}
        ebitda_m = {
            m: round(gross_margin_m[m] - fixed_exp_m[m] - var_exp_m[m], 2)
            for m in range(1, 13)
        }
        net_profit_m = {
            m: round(ebitda_m[m] + fin_res_m[m], 2) for m in range(1, 13)
        }

        gross_rev_tot = round(sum(gross_rev_m.values()), 2)
        cmv_tot = round(sum(cmv_m.values()), 2)
        gross_margin_tot = round(gross_rev_tot - cmv_tot, 2)
        fixed_exp_tot = round(sum(fixed_exp_m.values()), 2)
        var_exp_tot = round(sum(var_exp_m.values()), 2)
        ebitda_tot = round(gross_margin_tot - fixed_exp_tot - var_exp_tot, 2)
        fin_res_tot = round(sum(fin_res_m.values()), 2)
        net_profit_tot = round(ebitda_tot + fin_res_tot, 2)

        # 7. Preenche a Análise Vertical (% da receita bruta do mês e do ano)
        def calc_pct(val: float, base: float) -> float:
            if base and abs(base) > 0.001:
                return round((val / base) * 100.0, 2)
            return 0.0

        for group_type in group_rows_map:
            for row in group_rows_map[group_type]:
                row.monthly_percentages = {
                    m: calc_pct(row.monthly_amounts[m], gross_rev_m[m]) for m in range(1, 13)
                }
                row.total_percentage = calc_pct(row.total_amount, gross_rev_tot)

        # 8. Monta as linhas de subtotais e blocos
        all_rows: List[DRERowData] = []
        groups_list: List[DREGroupData] = []

        for group_type, title, subtotal_name in groups_config:
            subtotal_monthly = {
                m: round(group_monthly_totals[group_type][m], 2) for m in range(1, 13)
            }
            subtotal_tot = round(sum(subtotal_monthly.values()), 2)
            subtotal_avg = round(subtotal_tot / 12.0, 2)
            subtotal_pcts = {
                m: calc_pct(subtotal_monthly[m], gross_rev_m[m]) for m in range(1, 13)
            }
            subtotal_tot_pct = calc_pct(subtotal_tot, gross_rev_tot)

            sub_row = DRERowData(
                id=f"subtotal-{group_type}",
                account_id=None,
                name=title,
                code=None,
                group_type=group_type,
                is_system=True,
                is_header=False,
                is_subtotal=True,
                is_result=False,
                is_percentage_row=False,
                is_editable=False,
                display_order=0,
                monthly_amounts=subtotal_monthly,
                monthly_percentages=subtotal_pcts,
                total_amount=subtotal_tot,
                total_percentage=subtotal_tot_pct,
                monthly_average=subtotal_avg,
            )

            group_obj = DREGroupData(
                group_type=group_type,
                title=title,
                subtotal_name=subtotal_name,
                rows=group_rows_map[group_type],
                subtotal_row=sub_row,
            )
            groups_list.append(group_obj)

            # Insere no all_rows ordenado
            all_rows.append(sub_row)
            all_rows.extend(group_rows_map[group_type])

            # Linhas intermediárias calculadas:
            if group_type == "cmv":
                # (=) RECEITA LÍQUIDA / MARGEM BRUTA
                margin_sub_row = DRERowData(
                    id="result-gross-margin",
                    account_id=None,
                    name="(=) RECEITA LÍQUIDA / MARGEM BRUTA",
                    code=None,
                    group_type="cmv",
                    is_system=True,
                    is_header=False,
                    is_subtotal=True,
                    is_result=True,
                    is_percentage_row=False,
                    is_editable=False,
                    display_order=999,
                    monthly_amounts=gross_margin_m,
                    monthly_percentages={
                        m: calc_pct(gross_margin_m[m], gross_rev_m[m]) for m in range(1, 13)
                    },
                    total_amount=gross_margin_tot,
                    total_percentage=calc_pct(gross_margin_tot, gross_rev_tot),
                    monthly_average=round(gross_margin_tot / 12.0, 2),
                )
                all_rows.append(margin_sub_row)

            elif group_type == "variable_expense":
                # (=) RESULTADO OPERACIONAL (EBITDA / LAJIDA)
                ebitda_sub_row = DRERowData(
                    id="result-ebitda",
                    account_id=None,
                    name="(=) RESULTADO OPERACIONAL (EBITDA)",
                    code=None,
                    group_type="variable_expense",
                    is_system=True,
                    is_header=False,
                    is_subtotal=True,
                    is_result=True,
                    is_percentage_row=False,
                    is_editable=False,
                    display_order=999,
                    monthly_amounts=ebitda_m,
                    monthly_percentages={
                        m: calc_pct(ebitda_m[m], gross_rev_m[m]) for m in range(1, 13)
                    },
                    total_amount=ebitda_tot,
                    total_percentage=calc_pct(ebitda_tot, gross_rev_tot),
                    monthly_average=round(ebitda_tot / 12.0, 2),
                )
                all_rows.append(ebitda_sub_row)

        # 9. Linha Final de Resultado e Percentual
        net_profit_row = DRERowData(
            id="result-net-profit",
            account_id=None,
            name="(=) RESULTADO LÍQUIDO DO EXERCÍCIO (LUCRO/PREJUÍZO)",
            code=None,
            group_type="result",
            is_system=True,
            is_header=False,
            is_subtotal=True,
            is_result=True,
            is_percentage_row=False,
            is_editable=False,
            display_order=9999,
            monthly_amounts=net_profit_m,
            monthly_percentages={
                m: calc_pct(net_profit_m[m], gross_rev_m[m]) for m in range(1, 13)
            },
            total_amount=net_profit_tot,
            total_percentage=calc_pct(net_profit_tot, gross_rev_tot),
            monthly_average=round(net_profit_tot / 12.0, 2),
        )
        all_rows.append(net_profit_row)

        # Linha Percentual do Resultado
        pct_monthly = {
            m: calc_pct(net_profit_m[m], gross_rev_m[m]) for m in range(1, 13)
        }
        net_profit_pct_row = DRERowData(
            id="result-net-profit-pct",
            account_id=None,
            name="RESULTADO EM % (MARGEM LÍQUIDA)",
            code=None,
            group_type="result",
            is_system=True,
            is_header=False,
            is_subtotal=True,
            is_result=True,
            is_percentage_row=True,
            is_editable=False,
            display_order=10000,
            monthly_amounts=pct_monthly,
            monthly_percentages=pct_monthly,
            total_amount=calc_pct(net_profit_tot, gross_rev_tot),
            total_percentage=calc_pct(net_profit_tot, gross_rev_tot),
            monthly_average=calc_pct(net_profit_tot, gross_rev_tot),
        )
        all_rows.append(net_profit_pct_row)

        summary = DRESummary(
            gross_revenue_total=gross_rev_tot,
            cmv_total=cmv_tot,
            gross_margin_total=gross_margin_tot,
            gross_margin_pct=calc_pct(gross_margin_tot, gross_rev_tot),
            fixed_expenses_total=fixed_exp_tot,
            variable_expenses_total=var_exp_tot,
            ebitda_total=ebitda_tot,
            ebitda_pct=calc_pct(ebitda_tot, gross_rev_tot),
            financial_result_total=fin_res_tot,
            net_profit_total=net_profit_tot,
            net_margin_pct=calc_pct(net_profit_tot, gross_rev_tot),
        )

        return DREReportResponse(
            year=year,
            months=list(range(1, 13)),
            groups=groups_list,
            all_rows=all_rows,
            summary=summary,
        )

    def upsert_entry(
        self,
        db: Session,
        tenant_id: int,
        account_id: int,
        year: int,
        month: int,
        amount: float,
        notes: Optional[str] = None,
        user_id: Optional[int] = None,
    ) -> DREEntry:
        return self.repo.upsert_entry(
            db, tenant_id, account_id, year, month, amount, notes, user_id
        )

    def batch_upsert_entries(
        self,
        db: Session,
        tenant_id: int,
        entries: List[DREEntryUpsert],
        user_id: Optional[int] = None,
    ) -> List[DREEntry]:
        return self.repo.batch_upsert_entries(
            db=db,
            tenant_id=tenant_id,
            entries=entries,
            user_id=user_id,
        )

    def replicate_entry(
        self,
        db: Session,
        tenant_id: int,
        data: DREEntryReplicate,
        user_id: Optional[int] = None,
    ) -> List[DREEntry]:
        return self.repo.replicate_entry(
            db,
            tenant_id=tenant_id,
            account_id=data.account_id,
            year=data.competence_year,
            start_month=data.start_month,
            end_month=data.end_month,
            amount=data.amount,
            notes=data.notes,
            user_id=user_id,
        )

    def list_accounts(self, db: Session, tenant_id: int) -> List[DREAccount]:
        return self.repo.get_accounts(db, tenant_id, active_only=False)

    def create_account(
        self, db: Session, tenant_id: int, data: DREAccountCreate
    ) -> DREAccount:
        return self.repo.create_account(db, tenant_id, data)

    def update_account(
        self, db: Session, tenant_id: int, account_id: int, data: DREAccountUpdate
    ) -> Optional[DREAccount]:
        return self.repo.update_account(db, tenant_id, account_id, data)

    def delete_account(self, db: Session, tenant_id: int, account_id: int) -> bool:
        return self.repo.delete_account(db, tenant_id, account_id)

    # ── EXPORTAÇÃO EXCEL (.XLSX) PROFISSIONAL ──────────────────────────────

    def export_dre_excel(self, db: Session, tenant_id: int, year: int) -> io.BytesIO:
        report = self.get_dre_report(db, tenant_id, year)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"DRE {year}"

        # Paleta de estilos inspirada na planilha do cliente
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        subtotal_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")  # Dourado âmbar
        subtotal_font = Font(name="Calibri", size=11, bold=True, color="000000")

        result_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Âmbar suave
        result_font = Font(name="Calibri", size=11, bold=True, color="92400E")

        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        month_names = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

        # Cabeçalho do Relatório
        ws.merge_cells("A1:P1")
        title_cell = ws["A1"]
        title_cell.value = f"RELATÓRIO DRE - DEMONSTRATIVO DE RESULTADO DO EXERCÍCIO ({year})"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[3].height = 24

        # Linha de Colunas
        headers = ["MÊS / ANO", f"TOTAL {year}", "MÉDIA MENSAL"] + [f"{m}/{str(year)[2:]}" for m in month_names]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            cell.border = thin_border

        current_row = 4
        currency_format = 'R$ #,##0.00;[Red](R$ #,##0.00);"-"'
        pct_format = '0.00%'

        for row_data in report.all_rows:
            ws.row_dimensions[current_row].height = 20
            is_sub = row_data.is_subtotal or row_data.is_result
            is_pct = row_data.is_percentage_row

            # Coluna Nome
            name_cell = ws.cell(row=current_row, column=1, value=("   " if not is_sub else "") + row_data.name)
            name_cell.border = thin_border

            # Formatação visual de linha
            if is_sub:
                if row_data.id.startswith("subtotal-"):
                    row_fill = subtotal_fill
                    row_font = subtotal_font
                else:
                    row_fill = result_fill
                    row_font = result_font
            else:
                row_fill = PatternFill(fill_type=None)
                row_font = Font(name="Calibri", size=10)

            name_cell.fill = row_fill
            name_cell.font = row_font

            # Coluna Total
            tot_val = row_data.total_amount if not is_pct else (row_data.total_amount / 100.0)
            tot_cell = ws.cell(row=current_row, column=2, value=tot_val)
            tot_cell.number_format = pct_format if is_pct else currency_format
            tot_cell.fill = row_fill
            tot_cell.font = row_font
            tot_cell.alignment = Alignment(horizontal="right", vertical="center")
            tot_cell.border = thin_border

            # Coluna Média
            avg_val = row_data.monthly_average if not is_pct else (row_data.monthly_average / 100.0)
            avg_cell = ws.cell(row=current_row, column=3, value=avg_val)
            avg_cell.number_format = pct_format if is_pct else currency_format
            avg_cell.fill = row_fill
            avg_cell.font = row_font
            avg_cell.alignment = Alignment(horizontal="right", vertical="center")
            avg_cell.border = thin_border

            # Colunas Meses (1-12)
            for m in range(1, 13):
                col_i = 3 + m
                m_val = row_data.monthly_amounts.get(m, 0.0)
                if is_pct:
                    m_val = m_val / 100.0

                m_cell = ws.cell(row=current_row, column=col_i, value=m_val)
                m_cell.number_format = pct_format if is_pct else currency_format
                m_cell.fill = row_fill
                m_cell.font = row_font
                m_cell.alignment = Alignment(horizontal="right", vertical="center")
                m_cell.border = thin_border

            current_row += 1

        # Ajuste de largura das colunas
        ws.column_dimensions["A"].width = 44
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 15
        for m in range(1, 13):
            col_letter = get_column_letter(3 + m)
            ws.column_dimensions[col_letter].width = 14

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
