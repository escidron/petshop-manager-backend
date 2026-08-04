from sqlalchemy.orm import Session
from fastapi import HTTPException, status
from .repository import PackageRepository
from .schemas import PackageCreate, PackageUpdate

class PackageService:
    def __init__(self):
        self.repository = PackageRepository()

    def create_package(self, db: Session, tenant_id: int, data: PackageCreate):
        return self.repository.create(db, tenant_id, data)

    def list_packages(self, db: Session, tenant_id: int):
        return self.repository.list(db, tenant_id)

    def get_package(self, db: Session, tenant_id: int, package_id: int):
        package = self.repository.get_by_id(db, tenant_id, package_id)
        if not package:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Pacote não encontrado",
            )
        return package

    def update_package(self, db: Session, tenant_id: int, package_id: int, data: PackageUpdate):
        package = self.get_package(db, tenant_id, package_id)
        return self.repository.update(db, package, data)

    def delete_package(self, db: Session, tenant_id: int, package_id: int):
        package = self.get_package(db, tenant_id, package_id)
        self.repository.delete(db, package)

    def export_to_excel(self, db: Session, tenant_id: int) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pacotes"
        
        headers = [
            "ID", "Nome", "Descrição", "Preço", "Validade (dias)", "Itens", "Ativo"
        ]
        ws.append(headers)
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        packages = self.repository.list(db, tenant_id)
        
        for pkg in packages:
            items_str = []
            for item in pkg.items:
                if item.service:
                    items_str.append(f"{item.quantity}x {item.service.name}")
                elif item.product:
                    items_str.append(f"{item.quantity}x {item.product.name}")
                    
            ws.append([
                pkg.id, pkg.name, pkg.description or "", 
                float(pkg.price_cents) / 100 if pkg.price_cents else 0.0, 
                pkg.validity_days or "",
                ", ".join(items_str),
                "Sim" if pkg.is_active else "Não"
            ])
                
        out = BytesIO()
        wb.save(out)
        return out.getvalue()
