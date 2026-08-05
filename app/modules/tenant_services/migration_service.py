import io
import openpyxl
from io import BytesIO
from typing import Dict, List, Any, Optional

def clean_str(val: Any) -> Optional[str]:
    if val is None or str(val).strip() == "":
        return None
    val_str = str(val).strip()
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return val_str

def normalize_porte(val: Any) -> Optional[str]:
    if not val:
        return None
    v = str(val).strip().lower()
    if v in ["pp", "mini", "micro"]:
        return "PP"
    elif v in ["p", "pequeno", "peq"]:
        return "P"
    elif v in ["m", "medio", "médio"]:
        return "M"
    elif v in ["g", "grande"]:
        return "G"
    elif v in ["gg", "gigante", "extra grande"]:
        return "GG"
    if "(pp)" in v or "mini" in v: return "PP"
    if "(p)" in v or "pequeno" in v: return "P"
    if "(m)" in v or "médio" in v or "medio" in v: return "M"
    if "(g)" in v or "grande" in v: return "G"
    if "(gg)" in v or "gigante" in v: return "GG"
    return None


def parse_num(val: Any) -> Optional[float]:
    if val is None or str(val).strip() == "":
        return None
    try:
        val_str = str(val).strip().replace("R$", "").replace("r$", "").replace(" ", "")
        if "," in val_str and "." in val_str:
            val_str = val_str.replace(".", "").replace(",", ".")
        elif "," in val_str:
            val_str = val_str.replace(",", ".")
        return float(val_str)
    except Exception:
        return None

class ServiceMigrationService:
    def inspect_excel_headers(self, file_content: bytes) -> List[str]:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        return [str(h).strip() for h in rows[0] if h is not None and str(h).strip() != ""]

    def parse_excel_rows(self, file_content: bytes) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        parsed = []
        for r_vals in rows[1:]:
            r_dict = {}
            has_val = False
            for h, v in zip(headers, r_vals):
                if h:
                    r_dict[h] = v
                    if v is not None and str(v).strip() != "":
                        has_val = True
            if has_val:
                parsed.append(r_dict)
        return parsed

    def map_file(
        self,
        file_content: bytes,
        column_mapping: Dict[str, str],
        default_values: Dict[str, str] = None,
    ) -> List[Dict[str, Any]]:
        default_values = default_values or {}
        rows = self.parse_excel_rows(file_content)
        transformed = []
        for r in rows:
            item = {}
            for target_field, source_col in column_mapping.items():
                if source_col and source_col in r:
                    val = r[source_col]
                    if target_field in ("preco", "duracao_minutos"):
                        val = parse_num(val)
                    elif target_field == "porte":
                        val = normalize_porte(val)
                    else:
                        val = clean_str(val)
                    item[target_field] = val
                else:
                    item[target_field] = default_values.get(target_field)

            for k, def_val in default_values.items():
                if not item.get(k) and def_val:
                    item[k] = def_val

            if item.get("nome"):
                transformed.append(item)
        return transformed

    def generate_converted_excel(self, items: List[Dict[str, Any]]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Serviços"

        headers = [
            "servico *", "preco *", "especie", "porte",
            "pelagem", "duracao_minutos", "descricao"
        ]
        ws.append(headers)

        header_keys = [
            "nome", "preco", "especie", "porte",
            "pelagem", "duracao_minutos", "descricao"
        ]

        for item in items:
            ws.append([item.get(k, "") or "" for k in header_keys])

        out = BytesIO()
        wb.save(out)
        return out.getvalue()


